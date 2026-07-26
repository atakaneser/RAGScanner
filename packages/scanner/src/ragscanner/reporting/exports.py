"""Standalone dashboard report exports for HTML, XLSX, and PDF."""

from __future__ import annotations

import html
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Literal, cast

import reportlab
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    KeepTogether,
    LongTable,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

from ragscanner.reporting.localization import (
    localize_coverage_reason,
    localize_rule_field,
)
from ragscanner.reporting.models import ReportDocument, ReportFinding

ReportExportFormat = Literal["html", "xlsx", "pdf"]
SUPPORTED_REPORT_EXPORTS: tuple[ReportExportFormat, ...] = ("html", "xlsx", "pdf")
_LOCALES = frozenset({"en", "tr", "de", "fr", "zh-CN", "it"})
_MAX_CELL_LENGTH = 32_000
_PDF_OCCURRENCES_PER_GROUP = 20


@dataclass(frozen=True)
class ReportExport:
    content: bytes
    media_type: str
    extension: str


_TRANSLATIONS: dict[str, dict[str, str]] = {
    "tr": {
        "RAGScanner report": "RAGScanner raporu",
        "Executive summary": "Yönetici özeti",
        "Generated": "Oluşturulma",
        "Source": "Kaynak",
        "Status": "Durum",
        "Overall score": "Genel puan",
        "Security score": "Güvenlik puanı",
        "Content quality": "İçerik kalitesi",
        "Efficiency": "Verimlilik",
        "Not assessed": "Değerlendirilmedi",
        "Files discovered": "Bulunan dosyalar",
        "Files processed": "İşlenen dosyalar",
        "Files skipped": "Atlanan dosyalar",
        "Findings": "Bulgular",
        "Severity distribution": "Önem dağılımı",
        "AI-assisted analysis": "AI destekli analiz",
        "Priority actions": "Öncelikli eylemler",
        "Questions for review": "İnceleme soruları",
        "Verification steps": "Doğrulama adımları",
        "Limitations": "Sınırlamalar",
        "AI analysis unavailable": "AI analizi kullanılamıyor",
        "Source location": "Kaynak konumu",
        "Page": "Sayfa",
        "Line": "Satır",
        "Impact": "Etki",
        "Evidence": "Kanıt",
        "Problematic text": "Sorunlu metin",
        "Recommendation": "Öneri",
        "AI-assisted fix": "AI destekli çözüm",
        "Coverage": "Kapsam",
        "Area": "Alan",
        "Reason": "Neden",
        "Ingestion issues": "Veri alımı sorunları",
        "Path": "Yol",
        "Stage": "Aşama",
        "Issue": "Sorun",
        "Remediation": "Düzeltme",
        "No findings recorded.": "Kayıtlı bulgu yok.",
        "No ingestion issues recorded.": "Kayıtlı veri alımı sorunu yok.",
        "Methodology": "Metodoloji",
        "Report details": "Rapor ayrıntıları",
        "Rule": "Kural",
        "Title": "Başlık",
        "Category": "Kategori",
        "Severity": "Önem",
        "Confidence": "Güven",
        "File": "Dosya",
        "Summary": "Özet",
        "Scores": "Puanlar",
        "AI Analysis": "AI Analizi",
        "Critical": "Kritik",
        "High": "Yüksek",
        "Medium": "Orta",
        "Low": "Düşük",
        "Info": "Bilgi",
        "Occurrences": "Tekrarlar",
        "{count} more occurrences omitted; use HTML or Excel for the complete finding list.": "{count} tekrar daha PDF özetinden çıkarıldı; tam bulgu listesi için HTML veya Excel kullanın.",
        "Empty Chunk": "Boş chunk",
        "Punctuation Only Chunk": "Yalnızca noktalama içeren chunk",
        "Excessive Overlap": "Aşırı bindirme",
        "Unrelated Heading Branches": "İlgisiz başlık dalları",
        "Exact normalized-content duplicate group": "Tam normalleştirilmiş içerik tekrar grubu",
        "Poor chunk quality can reduce retrieval precision, waste context, or hide source structure.": "Düşük chunk kalitesi arama doğruluğunu azaltabilir, bağlamı israf edebilir veya kaynak yapısını gizleyebilir.",
        "Redundant indexed content can waste storage and bias retrieval.": "Gereksiz yinelenen indeks içeriği depolamayı israf edebilir ve aramayı yanlı hale getirebilir.",
        "Review the chunk and adjust deterministic chunking configuration if appropriate.": "Chunk'ı inceleyin ve gerekiyorsa deterministik parçalama ayarını düzeltin.",
        "Reduce bounded overlap without crossing unrelated structural boundaries.": "İlgisiz yapısal sınırları aşmadan bindirmeyi azaltın.",
        "Review the group and keep one canonical item; do not delete automatically.": "Grubu inceleyin ve tek bir kanonik öğe bırakın; otomatik olarak silmeyin.",
        "AI-generated analysis is advisory. Verify it against the deterministic findings and underlying evidence before acting.": "AI tarafından üretilen analiz tavsiye niteliğindedir. İşlem yapmadan önce deterministik bulgular ve dayanak kanıtlarla doğrulayın.",
        "completed": "tamamlandı",
        "completed_with_warnings": "uyarılarla tamamlandı",
        "assessed": "değerlendirildi",
        "not_assessed": "değerlendirilmedi",
        "Generated locally by RAGScanner. No external assets or network requests.": "RAGScanner tarafından yerel olarak oluşturuldu. Harici varlık veya ağ isteği içermez.",
    },
    "de": {
        "RAGScanner report": "RAGScanner-Bericht",
        "Executive summary": "Zusammenfassung",
        "Generated": "Erstellt",
        "Source": "Quelle",
        "Status": "Status",
        "Overall score": "Gesamtbewertung",
        "Security score": "Sicherheitsbewertung",
        "Content quality": "Inhaltsqualität",
        "Efficiency": "Effizienz",
        "Not assessed": "Nicht bewertet",
        "Files discovered": "Gefundene Dateien",
        "Files processed": "Verarbeitete Dateien",
        "Files skipped": "Übersprungene Dateien",
        "Findings": "Befunde",
        "Severity distribution": "Schweregradverteilung",
        "AI-assisted analysis": "KI-gestützte Analyse",
        "Priority actions": "Priorisierte Maßnahmen",
        "Questions for review": "Prüffragen",
        "Verification steps": "Prüfschritte",
        "Limitations": "Einschränkungen",
        "AI analysis unavailable": "KI-Analyse nicht verfügbar",
        "Source location": "Quellposition",
        "Page": "Seite",
        "Line": "Zeile",
        "Impact": "Auswirkung",
        "Evidence": "Nachweis",
        "Problematic text": "Problematischer Text",
        "Recommendation": "Empfehlung",
        "AI-assisted fix": "KI-gestützte Behebung",
        "Coverage": "Abdeckung",
        "Area": "Bereich",
        "Reason": "Grund",
        "Ingestion issues": "Aufnahmeprobleme",
        "Path": "Pfad",
        "Stage": "Phase",
        "Issue": "Problem",
        "Remediation": "Behebung",
        "No findings recorded.": "Keine Befunde erfasst.",
        "No ingestion issues recorded.": "Keine Aufnahmeprobleme erfasst.",
        "Methodology": "Methodik",
        "Report details": "Berichtsdetails",
        "Rule": "Regel",
        "Title": "Titel",
        "Category": "Kategorie",
        "Severity": "Schweregrad",
        "Confidence": "Konfidenz",
        "File": "Datei",
        "Summary": "Übersicht",
        "Scores": "Bewertungen",
        "AI Analysis": "KI-Analyse",
        "Critical": "Kritisch",
        "High": "Hoch",
        "Medium": "Mittel",
        "Low": "Niedrig",
        "Info": "Info",
        "Occurrences": "Vorkommen",
        "{count} more occurrences omitted; use HTML or Excel for the complete finding list.": "{count} weitere Vorkommen wurden in der PDF-Zusammenfassung ausgelassen; die vollständige Liste ist in HTML oder Excel verfügbar.",
        "Empty Chunk": "Leerer Chunk",
        "Punctuation Only Chunk": "Chunk nur mit Satzzeichen",
        "Excessive Overlap": "Übermäßige Überlappung",
        "Unrelated Heading Branches": "Unabhängige Überschriftenzweige",
        "Exact normalized-content duplicate group": "Gruppe exakt normalisierter Inhaltsduplikate",
        "Poor chunk quality can reduce retrieval precision, waste context, or hide source structure.": "Schlechte Chunk-Qualität kann die Abrufpräzision verringern, Kontext verschwenden oder die Quellstruktur verbergen.",
        "Redundant indexed content can waste storage and bias retrieval.": "Redundanter indexierter Inhalt kann Speicher verschwenden und den Abruf verzerren.",
        "Review the chunk and adjust deterministic chunking configuration if appropriate.": "Prüfen Sie den Chunk und passen Sie bei Bedarf die deterministische Chunk-Konfiguration an.",
        "Reduce bounded overlap without crossing unrelated structural boundaries.": "Reduzieren Sie die begrenzte Überlappung, ohne unabhängige Strukturgrenzen zu überschreiten.",
        "Review the group and keep one canonical item; do not delete automatically.": "Prüfen Sie die Gruppe und behalten Sie ein kanonisches Element; nicht automatisch löschen.",
        "AI-generated analysis is advisory. Verify it against the deterministic findings and underlying evidence before acting.": "KI-generierte Analysen sind Hinweise. Prüfen Sie sie vor Maßnahmen anhand der deterministischen Befunde und Belege.",
        "completed": "abgeschlossen",
        "completed_with_warnings": "mit Warnungen abgeschlossen",
        "assessed": "bewertet",
        "not_assessed": "nicht bewertet",
        "Generated locally by RAGScanner. No external assets or network requests.": "Lokal von RAGScanner erstellt. Keine externen Ressourcen oder Netzwerkanfragen.",
    },
    "fr": {
        "RAGScanner report": "Rapport RAGScanner",
        "Executive summary": "Résumé exécutif",
        "Generated": "Généré",
        "Source": "Source",
        "Status": "État",
        "Overall score": "Score global",
        "Security score": "Score de sécurité",
        "Content quality": "Qualité du contenu",
        "Efficiency": "Efficacité",
        "Not assessed": "Non évalué",
        "Files discovered": "Fichiers trouvés",
        "Files processed": "Fichiers traités",
        "Files skipped": "Fichiers ignorés",
        "Findings": "Constats",
        "Severity distribution": "Répartition de la sévérité",
        "AI-assisted analysis": "Analyse assistée par IA",
        "Priority actions": "Actions prioritaires",
        "Questions for review": "Questions de contrôle",
        "Verification steps": "Étapes de vérification",
        "Limitations": "Limites",
        "AI analysis unavailable": "Analyse IA indisponible",
        "Source location": "Emplacement source",
        "Page": "Page",
        "Line": "Ligne",
        "Impact": "Impact",
        "Evidence": "Preuve",
        "Problematic text": "Texte problématique",
        "Recommendation": "Recommandation",
        "AI-assisted fix": "Correction assistée par IA",
        "Coverage": "Couverture",
        "Area": "Domaine",
        "Reason": "Raison",
        "Ingestion issues": "Problèmes d’ingestion",
        "Path": "Chemin",
        "Stage": "Étape",
        "Issue": "Problème",
        "Remediation": "Correction",
        "No findings recorded.": "Aucun constat enregistré.",
        "No ingestion issues recorded.": "Aucun problème d’ingestion enregistré.",
        "Methodology": "Méthodologie",
        "Report details": "Détails du rapport",
        "Rule": "Règle",
        "Title": "Titre",
        "Category": "Catégorie",
        "Severity": "Sévérité",
        "Confidence": "Confiance",
        "File": "Fichier",
        "Summary": "Résumé",
        "Scores": "Scores",
        "AI Analysis": "Analyse IA",
        "Critical": "Critique",
        "High": "Élevée",
        "Medium": "Moyenne",
        "Low": "Faible",
        "Info": "Info",
        "Occurrences": "Occurrences",
        "{count} more occurrences omitted; use HTML or Excel for the complete finding list.": "{count} occurrences supplémentaires ont été omises du résumé PDF ; utilisez HTML ou Excel pour la liste complète.",
        "Empty Chunk": "Fragment vide",
        "Punctuation Only Chunk": "Fragment composé uniquement de ponctuation",
        "Excessive Overlap": "Chevauchement excessif",
        "Unrelated Heading Branches": "Branches de titres sans rapport",
        "Exact normalized-content duplicate group": "Groupe de contenu normalisé exactement dupliqué",
        "Poor chunk quality can reduce retrieval precision, waste context, or hide source structure.": "Une mauvaise qualité de fragment peut réduire la précision de recherche, gaspiller le contexte ou masquer la structure source.",
        "Redundant indexed content can waste storage and bias retrieval.": "Le contenu indexé redondant peut gaspiller du stockage et biaiser la recherche.",
        "Review the chunk and adjust deterministic chunking configuration if appropriate.": "Examinez le fragment et ajustez la configuration déterministe si nécessaire.",
        "Reduce bounded overlap without crossing unrelated structural boundaries.": "Réduisez le chevauchement limité sans franchir de limites structurelles sans rapport.",
        "Review the group and keep one canonical item; do not delete automatically.": "Examinez le groupe et conservez un élément canonique ; ne supprimez rien automatiquement.",
        "AI-generated analysis is advisory. Verify it against the deterministic findings and underlying evidence before acting.": "L’analyse générée par l’IA est consultative. Vérifiez-la avec les constats déterministes et les preuves avant d’agir.",
        "completed": "terminé",
        "completed_with_warnings": "terminé avec avertissements",
        "assessed": "évalué",
        "not_assessed": "non évalué",
        "Generated locally by RAGScanner. No external assets or network requests.": "Généré localement par RAGScanner. Aucune ressource externe ni requête réseau.",
    },
    "zh-CN": {
        "RAGScanner report": "RAGScanner 报告",
        "Executive summary": "执行摘要",
        "Generated": "生成时间",
        "Source": "数据源",
        "Status": "状态",
        "Overall score": "总分",
        "Security score": "安全分数",
        "Content quality": "内容质量",
        "Efficiency": "效率",
        "Not assessed": "未评估",
        "Files discovered": "发现的文件",
        "Files processed": "已处理文件",
        "Files skipped": "已跳过文件",
        "Findings": "发现",
        "Severity distribution": "严重程度分布",
        "AI-assisted analysis": "AI 辅助分析",
        "Priority actions": "优先操作",
        "Questions for review": "审核问题",
        "Verification steps": "验证步骤",
        "Limitations": "局限性",
        "AI analysis unavailable": "AI 分析不可用",
        "Source location": "来源位置",
        "Page": "页",
        "Line": "行",
        "Impact": "影响",
        "Evidence": "证据",
        "Problematic text": "问题文本",
        "Recommendation": "建议",
        "AI-assisted fix": "AI 辅助修复",
        "Coverage": "覆盖率",
        "Area": "领域",
        "Reason": "原因",
        "Ingestion issues": "摄取问题",
        "Path": "路径",
        "Stage": "阶段",
        "Issue": "问题",
        "Remediation": "修复",
        "No findings recorded.": "没有记录任何发现。",
        "No ingestion issues recorded.": "没有记录摄取问题。",
        "Methodology": "方法",
        "Report details": "报告详情",
        "Rule": "规则",
        "Title": "标题",
        "Category": "类别",
        "Severity": "严重程度",
        "Confidence": "置信度",
        "File": "文件",
        "Summary": "摘要",
        "Scores": "分数",
        "AI Analysis": "AI 分析",
        "Critical": "严重",
        "High": "高",
        "Medium": "中",
        "Low": "低",
        "Info": "信息",
        "Occurrences": "出现位置",
        "{count} more occurrences omitted; use HTML or Excel for the complete finding list.": "PDF 摘要省略了另外 {count} 个位置；请使用 HTML 或 Excel 查看完整发现列表。",
        "Empty Chunk": "空分块",
        "Punctuation Only Chunk": "仅含标点的分块",
        "Excessive Overlap": "重叠过多",
        "Unrelated Heading Branches": "不相关的标题分支",
        "Exact normalized-content duplicate group": "完全相同的规范化内容重复组",
        "Poor chunk quality can reduce retrieval precision, waste context, or hide source structure.": "分块质量不佳可能降低检索精度、浪费上下文或隐藏来源结构。",
        "Redundant indexed content can waste storage and bias retrieval.": "冗余索引内容可能浪费存储空间并使检索产生偏差。",
        "Review the chunk and adjust deterministic chunking configuration if appropriate.": "检查该分块，并在适当时调整确定性分块配置。",
        "Reduce bounded overlap without crossing unrelated structural boundaries.": "在不跨越无关结构边界的前提下减少受限重叠。",
        "Review the group and keep one canonical item; do not delete automatically.": "检查该组并保留一个规范项；不要自动删除。",
        "AI-generated analysis is advisory. Verify it against the deterministic findings and underlying evidence before acting.": "AI 生成的分析仅供参考。采取行动前，请根据确定性发现和基础证据进行核实。",
        "completed": "已完成",
        "completed_with_warnings": "已完成但有警告",
        "assessed": "已评估",
        "not_assessed": "未评估",
        "Generated locally by RAGScanner. No external assets or network requests.": "由 RAGScanner 在本地生成，不包含外部资源或网络请求。",
    },
    "it": {
        "RAGScanner report": "Rapporto RAGScanner",
        "Executive summary": "Riepilogo esecutivo",
        "Generated": "Generato",
        "Source": "Origine",
        "Status": "Stato",
        "Overall score": "Punteggio complessivo",
        "Security score": "Punteggio sicurezza",
        "Content quality": "Qualità del contenuto",
        "Efficiency": "Efficienza",
        "Not assessed": "Non valutato",
        "Files discovered": "File trovati",
        "Files processed": "File elaborati",
        "Files skipped": "File ignorati",
        "Findings": "Risultati",
        "Severity distribution": "Distribuzione gravità",
        "AI-assisted analysis": "Analisi assistita dall’AI",
        "Priority actions": "Azioni prioritarie",
        "Questions for review": "Domande di revisione",
        "Verification steps": "Passaggi di verifica",
        "Limitations": "Limitazioni",
        "AI analysis unavailable": "Analisi AI non disponibile",
        "Source location": "Posizione origine",
        "Page": "Pagina",
        "Line": "Riga",
        "Impact": "Impatto",
        "Evidence": "Prova",
        "Problematic text": "Testo problematico",
        "Recommendation": "Raccomandazione",
        "AI-assisted fix": "Correzione assistita dall’AI",
        "Coverage": "Copertura",
        "Area": "Area",
        "Reason": "Motivo",
        "Ingestion issues": "Problemi di acquisizione",
        "Path": "Percorso",
        "Stage": "Fase",
        "Issue": "Problema",
        "Remediation": "Correzione",
        "No findings recorded.": "Nessun risultato registrato.",
        "No ingestion issues recorded.": "Nessun problema di acquisizione registrato.",
        "Methodology": "Metodologia",
        "Report details": "Dettagli rapporto",
        "Rule": "Regola",
        "Title": "Titolo",
        "Category": "Categoria",
        "Severity": "Gravità",
        "Confidence": "Confidenza",
        "File": "File",
        "Summary": "Riepilogo",
        "Scores": "Punteggi",
        "AI Analysis": "Analisi AI",
        "Critical": "Critica",
        "High": "Alta",
        "Medium": "Media",
        "Low": "Bassa",
        "Info": "Info",
        "Occurrences": "Occorrenze",
        "{count} more occurrences omitted; use HTML or Excel for the complete finding list.": "Altre {count} occorrenze sono state omesse dal riepilogo PDF; usa HTML o Excel per l'elenco completo.",
        "Empty Chunk": "Chunk vuoto",
        "Punctuation Only Chunk": "Chunk di sola punteggiatura",
        "Excessive Overlap": "Sovrapposizione eccessiva",
        "Unrelated Heading Branches": "Rami di intestazione non correlati",
        "Exact normalized-content duplicate group": "Gruppo di contenuti normalizzati duplicati esatti",
        "Poor chunk quality can reduce retrieval precision, waste context, or hide source structure.": "Una scarsa qualità dei chunk può ridurre la precisione del recupero, sprecare contesto o nascondere la struttura della fonte.",
        "Redundant indexed content can waste storage and bias retrieval.": "Il contenuto indicizzato ridondante può sprecare spazio e distorcere il recupero.",
        "Review the chunk and adjust deterministic chunking configuration if appropriate.": "Esamina il chunk e modifica la configurazione deterministica se necessario.",
        "Reduce bounded overlap without crossing unrelated structural boundaries.": "Riduci la sovrapposizione limitata senza attraversare confini strutturali non correlati.",
        "Review the group and keep one canonical item; do not delete automatically.": "Esamina il gruppo e conserva un elemento canonico; non eliminare automaticamente.",
        "AI-generated analysis is advisory. Verify it against the deterministic findings and underlying evidence before acting.": "L’analisi generata dall’AI è consultiva. Verificala con i risultati deterministici e le prove prima di agire.",
        "completed": "completato",
        "completed_with_warnings": "completato con avvisi",
        "assessed": "valutato",
        "not_assessed": "non valutato",
        "Generated locally by RAGScanner. No external assets or network requests.": "Generato localmente da RAGScanner. Nessuna risorsa esterna o richiesta di rete.",
    },
}

_RAG_TRANSLATIONS = {
    "tr": {
        "RAG configuration advice": "RAG yapılandırma önerisi",
        "RAG Configuration": "RAG Yapılandırması",
        "Workload profile": "İş yükü profili",
        "Recommended chunk range": "Önerilen chunk aralığı",
        "Overlap": "Bindirme",
        "Retrieval top-k": "Getirme top-k",
        "Actions": "Eylemler",
        "Validation metrics": "Doğrulama metrikleri",
        "Why this range?": "Neden bu aralık?",
        "Recommended why": "Önerinin gerekçesi",
        "tokens": "token",
        "target": "hedef",
        "Starting point only; validate with representative queries before production use.": "Yalnızca başlangıç noktasıdır; üretimden önce temsili sorgularla doğrulayın.",
        "fact_lookup": "olgusal arama",
        "general_qa": "genel soru-cevap",
        "policy_procedure": "politika ve prosedür",
        "long_context_research": "uzun bağlam araştırması",
        "code_assistant": "kod asistanı",
        "table_analytics": "tablo analizi",
        "structure_aware": "yapıya duyarlı",
    },
    "de": {
        "RAG configuration advice": "RAG-Konfigurationsempfehlung",
        "RAG Configuration": "RAG-Konfiguration",
        "Workload profile": "Arbeitslastprofil",
        "Recommended chunk range": "Empfohlener Chunk-Bereich",
        "Overlap": "Überlappung",
        "Retrieval top-k": "Abruf Top-k",
        "Actions": "Maßnahmen",
        "Validation metrics": "Validierungsmetriken",
        "Why this range?": "Warum dieser Bereich?",
        "Recommended why": "Begründung der Empfehlung",
        "tokens": "Token",
        "target": "Ziel",
        "Starting point only; validate with representative queries before production use.": "Nur ein Ausgangspunkt; vor dem Produktionseinsatz mit repräsentativen Abfragen validieren.",
        "fact_lookup": "Faktensuche",
        "general_qa": "allgemeine Fragen und Antworten",
        "policy_procedure": "Richtlinien und Verfahren",
        "long_context_research": "Langkontext-Recherche",
        "code_assistant": "Code-Assistent",
        "table_analytics": "Tabellenanalyse",
        "structure_aware": "strukturbewusst",
    },
    "fr": {
        "RAG configuration advice": "Conseil de configuration RAG",
        "RAG Configuration": "Configuration RAG",
        "Workload profile": "Profil de charge",
        "Recommended chunk range": "Plage de fragments recommandée",
        "Overlap": "Chevauchement",
        "Retrieval top-k": "Top-k de recherche",
        "Actions": "Actions",
        "Validation metrics": "Métriques de validation",
        "Why this range?": "Pourquoi cette plage ?",
        "Recommended why": "Justification de la recommandation",
        "tokens": "jetons",
        "target": "cible",
        "Starting point only; validate with representative queries before production use.": "Point de départ uniquement ; validez avec des requêtes représentatives avant la production.",
        "fact_lookup": "recherche factuelle",
        "general_qa": "questions-réponses générales",
        "policy_procedure": "politiques et procédures",
        "long_context_research": "recherche en contexte long",
        "code_assistant": "assistant de code",
        "table_analytics": "analyse de tableaux",
        "structure_aware": "sensible à la structure",
    },
    "zh-CN": {
        "RAG configuration advice": "RAG 配置建议",
        "RAG Configuration": "RAG 配置",
        "Workload profile": "工作负载配置",
        "Recommended chunk range": "建议分块范围",
        "Overlap": "重叠",
        "Retrieval top-k": "检索 top-k",
        "Actions": "操作",
        "Validation metrics": "验证指标",
        "Why this range?": "为何采用此范围？",
        "Recommended why": "建议理由",
        "tokens": "令牌",
        "target": "目标",
        "Starting point only; validate with representative queries before production use.": "仅作为起点；生产使用前请用代表性查询进行验证。",
        "fact_lookup": "事实检索",
        "general_qa": "通用问答",
        "policy_procedure": "政策与流程",
        "long_context_research": "长上下文研究",
        "code_assistant": "代码助手",
        "table_analytics": "表格分析",
        "structure_aware": "结构感知",
    },
    "it": {
        "RAG configuration advice": "Suggerimento di configurazione RAG",
        "RAG Configuration": "Configurazione RAG",
        "Workload profile": "Profilo del carico",
        "Recommended chunk range": "Intervallo di chunk consigliato",
        "Overlap": "Sovrapposizione",
        "Retrieval top-k": "Top-k di recupero",
        "Actions": "Azioni",
        "Validation metrics": "Metriche di validazione",
        "Why this range?": "Perché questo intervallo?",
        "Recommended why": "Motivo del suggerimento",
        "tokens": "token",
        "target": "obiettivo",
        "Starting point only; validate with representative queries before production use.": "Solo un punto di partenza; valida con query rappresentative prima dell'uso in produzione.",
        "fact_lookup": "ricerca fattuale",
        "general_qa": "domande e risposte generali",
        "policy_procedure": "politiche e procedure",
        "long_context_research": "ricerca a contesto lungo",
        "code_assistant": "assistente di codice",
        "table_analytics": "analisi di tabelle",
        "structure_aware": "sensibile alla struttura",
    },
}
for _rag_locale, _rag_catalog in _RAG_TRANSLATIONS.items():
    _TRANSLATIONS[_rag_locale].update(_rag_catalog)

_AI_SECTION_TRANSLATIONS: dict[str, dict[str, str]] = {
    "tr": {
        "Root cause analysis": "Kök neden analizi",
        "Score commentary": "Puan yorumu",
        "Coverage caveat": "Kapsam uyarısı",
        "Target": "Hedef",
        "Effort": "Efor",
        "Addresses": "Ele aldığı bulgular",
        "Expected effect": "Beklenen etki",
        "Informs": "Karara katkısı",
        "ingestion": "veri alımı",
        "chunking": "parçalama",
        "corpus": "doküman kümesi",
        "configuration": "yapılandırma",
        "confirmed": "doğrulandı",
        "likely": "olası",
        "low": "düşük",
        "medium": "orta",
        "high": "yüksek",
    },
    "de": {
        "Root cause analysis": "Ursachenanalyse",
        "Score commentary": "Bewertungskommentar",
        "Coverage caveat": "Abdeckungshinweis",
        "Target": "Ziel",
        "Effort": "Aufwand",
        "Addresses": "Behandelte Befunde",
        "Expected effect": "Erwartete Wirkung",
        "Informs": "Unterstützte Entscheidung",
        "ingestion": "Aufnahme",
        "chunking": "Chunking",
        "corpus": "Korpus",
        "configuration": "Konfiguration",
        "confirmed": "bestätigt",
        "likely": "wahrscheinlich",
        "low": "niedrig",
        "medium": "mittel",
        "high": "hoch",
    },
    "fr": {
        "Root cause analysis": "Analyse des causes profondes",
        "Score commentary": "Commentaire sur le score",
        "Coverage caveat": "Avertissement de couverture",
        "Target": "Cible",
        "Effort": "Effort",
        "Addresses": "Constats traités",
        "Expected effect": "Effet attendu",
        "Informs": "Décision éclairée",
        "ingestion": "ingestion",
        "chunking": "segmentation",
        "corpus": "corpus",
        "configuration": "configuration",
        "confirmed": "confirmé",
        "likely": "probable",
        "low": "faible",
        "medium": "moyen",
        "high": "élevé",
    },
    "zh-CN": {
        "Root cause analysis": "根因分析",
        "Score commentary": "分数说明",
        "Coverage caveat": "覆盖范围提示",
        "Target": "目标",
        "Effort": "工作量",
        "Addresses": "处理的发现",
        "Expected effect": "预期效果",
        "Informs": "辅助决策",
        "ingestion": "摄取",
        "chunking": "分块",
        "corpus": "语料库",
        "configuration": "配置",
        "confirmed": "已确认",
        "likely": "可能",
        "low": "低",
        "medium": "中",
        "high": "高",
    },
    "it": {
        "Root cause analysis": "Analisi delle cause principali",
        "Score commentary": "Commento sul punteggio",
        "Coverage caveat": "Avvertenza sulla copertura",
        "Target": "Obiettivo",
        "Effort": "Impegno",
        "Addresses": "Risultati affrontati",
        "Expected effect": "Effetto previsto",
        "Informs": "Decisione supportata",
        "ingestion": "acquisizione",
        "chunking": "suddivisione",
        "corpus": "corpus",
        "configuration": "configurazione",
        "confirmed": "confermato",
        "likely": "probabile",
        "low": "basso",
        "medium": "medio",
        "high": "alto",
    },
}
for _ai_locale, _ai_catalog in _AI_SECTION_TRANSLATIONS.items():
    _TRANSLATIONS[_ai_locale].update(_ai_catalog)

_REPORT_NARRATIVE_TRANSLATIONS: dict[str, dict[str, str]] = {
    "tr": {
        "Chunk quality": "Parça kalitesi",
        "Cross-document exact duplicates": "Belgeler arası tam yinelenen içerik",
        "Cross-document freshness": "Belgeler arası güncellik",
        "Cross-document near duplicates": "Belgeler arası yakın yinelenen içerik",
        "Static security": "Statik güvenlik",
        "Version conflicts": "Sürüm çakışmaları",
        "Within-document near duplicates": "Belge içindeki yakın yinelenen içerik",
        "Within-document repeated chunks": "Belge içindeki yinelenen parçalar",
        "Static security rules, normalized-content duplicate analysis, and chunk-quality heuristics": "Statik güvenlik kuralları, normalleştirilmiş içerik yineleme analizi ve parça kalitesi sezgisel kontrolleri",
        "Scores are RAGScanner product-defined and do not prove retrieval or answer quality.": "Puanlar RAGScanner ürün tanımlarıdır; getirim veya yanıt kalitesini kanıtlamaz.",
        "Retrieval quality, answer reliability, freshness, and RAG Rot were not assessed.": "Getirim kalitesi, yanıt güvenilirliği, güncellik ve RAG Rot değerlendirilmedi.",
    },
    "de": {
        "Chunk quality": "Chunk-Qualität",
        "Cross-document exact duplicates": "Dokumentübergreifende exakte Duplikate",
        "Cross-document freshness": "Dokumentübergreifende Aktualität",
        "Cross-document near duplicates": "Dokumentübergreifende ähnliche Duplikate",
        "Static security": "Statische Sicherheit",
        "Version conflicts": "Versionskonflikte",
        "Within-document near duplicates": "Ähnliche Duplikate im Dokument",
        "Within-document repeated chunks": "Wiederholte Chunks im Dokument",
        "Static security rules, normalized-content duplicate analysis, and chunk-quality heuristics": "Statische Sicherheitsregeln, Analyse normalisierter Duplikate und Heuristiken zur Chunk-Qualität",
        "Scores are RAGScanner product-defined and do not prove retrieval or answer quality.": "Die Bewertungen sind RAGScanner-Produktdefinitionen und belegen weder Abruf- noch Antwortqualität.",
        "Retrieval quality, answer reliability, freshness, and RAG Rot were not assessed.": "Abrufqualität, Antwortzuverlässigkeit, Aktualität und RAG Rot wurden nicht bewertet.",
    },
    "fr": {
        "Chunk quality": "Qualité des fragments",
        "Cross-document exact duplicates": "Doublons exacts entre documents",
        "Cross-document freshness": "Actualité entre documents",
        "Cross-document near duplicates": "Quasi-doublons entre documents",
        "Static security": "Sécurité statique",
        "Version conflicts": "Conflits de versions",
        "Within-document near duplicates": "Quasi-doublons dans le document",
        "Within-document repeated chunks": "Fragments répétés dans le document",
        "Static security rules, normalized-content duplicate analysis, and chunk-quality heuristics": "Règles de sécurité statiques, analyse des doublons normalisés et heuristiques de qualité des fragments",
        "Scores are RAGScanner product-defined and do not prove retrieval or answer quality.": "Les scores sont définis par RAGScanner et ne prouvent ni la qualité de recherche ni celle des réponses.",
        "Retrieval quality, answer reliability, freshness, and RAG Rot were not assessed.": "La qualité de recherche, la fiabilité des réponses, l’actualité et RAG Rot n’ont pas été évalués.",
    },
    "zh-CN": {
        "Chunk quality": "分块质量",
        "Cross-document exact duplicates": "跨文档完全重复",
        "Cross-document freshness": "跨文档时效性",
        "Cross-document near duplicates": "跨文档近似重复",
        "Static security": "静态安全",
        "Version conflicts": "版本冲突",
        "Within-document near duplicates": "文档内近似重复",
        "Within-document repeated chunks": "文档内重复分块",
        "Static security rules, normalized-content duplicate analysis, and chunk-quality heuristics": "静态安全规则、规范化内容重复分析和分块质量启发式检查",
        "Scores are RAGScanner product-defined and do not prove retrieval or answer quality.": "分数由 RAGScanner 产品定义，不能证明检索或回答质量。",
        "Retrieval quality, answer reliability, freshness, and RAG Rot were not assessed.": "未评估检索质量、回答可靠性、时效性和 RAG Rot。",
    },
    "it": {
        "Chunk quality": "Qualità dei chunk",
        "Cross-document exact duplicates": "Duplicati esatti tra documenti",
        "Cross-document freshness": "Aggiornamento tra documenti",
        "Cross-document near duplicates": "Quasi duplicati tra documenti",
        "Static security": "Sicurezza statica",
        "Version conflicts": "Conflitti di versione",
        "Within-document near duplicates": "Quasi duplicati nel documento",
        "Within-document repeated chunks": "Chunk ripetuti nel documento",
        "Static security rules, normalized-content duplicate analysis, and chunk-quality heuristics": "Regole di sicurezza statiche, analisi dei duplicati normalizzati ed euristiche sulla qualità dei chunk",
        "Scores are RAGScanner product-defined and do not prove retrieval or answer quality.": "I punteggi sono definiti da RAGScanner e non dimostrano la qualità del recupero o delle risposte.",
        "Retrieval quality, answer reliability, freshness, and RAG Rot were not assessed.": "Qualità del recupero, affidabilità delle risposte, aggiornamento e RAG Rot non sono stati valutati.",
    },
}
for _narrative_locale, _narrative_catalog in _REPORT_NARRATIVE_TRANSLATIONS.items():
    _TRANSLATIONS[_narrative_locale].update(_narrative_catalog)

_RAG_NARRATIVE_TRANSLATIONS: dict[str, dict[str, str]] = {
    "Observed chunks are more fragmented than this profile; merge adjacent semantic sections and re-evaluate retrieval.": {
        "tr": "Gözlenen chunk'lar bu profile göre fazla parçalı; komşu anlamsal bölümleri birleştirip getirmeyi yeniden değerlendirin.",
        "de": "Die beobachteten Chunks sind stärker fragmentiert; benachbarte semantische Abschnitte zusammenführen und den Abruf neu bewerten.",
        "fr": "Les fragments observés sont trop morcelés ; fusionnez les sections sémantiques voisines et réévaluez la recherche.",
        "zh-CN": "观测分块比该配置更碎；合并相邻语义段并重新评估检索。",
        "it": "I chunk osservati sono più frammentati; unisci le sezioni semantiche adiacenti e rivaluta il recupero.",
    },
    "Observed chunks are coarser than this profile; reduce size without splitting lists, tables, code, or procedures.": {
        "tr": "Gözlenen chunk'lar bu profile göre fazla geniş; listeleri, tabloları, kodu veya prosedürleri bölmeden boyutu azaltın.",
        "de": "Die beobachteten Chunks sind gröber; Größe reduzieren, ohne Listen, Tabellen, Code oder Verfahren zu trennen.",
        "fr": "Les fragments observés sont trop larges ; réduisez leur taille sans couper listes, tableaux, code ou procédures.",
        "zh-CN": "观测分块比该配置更粗；在不拆分列表、表格、代码或流程的情况下减小尺寸。",
        "it": "I chunk osservati sono più ampi; riduci la dimensione senza dividere elenchi, tabelle, codice o procedure.",
    },
    "Repair structural split findings before tuning token counts; boundary quality takes priority over a numeric target.": {
        "tr": "Token sayılarını ayarlamadan önce yapısal bölünme bulgularını düzeltin; sınır kalitesi sayısal hedeften önce gelir.",
        "de": "Strukturelle Teilungsbefunde vor der Token-Abstimmung beheben; Grenzqualität hat Vorrang.",
        "fr": "Corrigez les divisions structurelles avant d’ajuster les jetons ; la qualité des limites prime.",
        "zh-CN": "调整令牌数前先修复结构切分问题；边界质量优先于数值目标。",
        "it": "Correggi le divisioni strutturali prima dei token; la qualità dei confini ha priorità.",
    },
    "More than one quarter of assessed chunks are undersized; review heading-only and fragmented upstream chunks.": {
        "tr": "Değerlendirilen chunk'ların dörtte birinden fazlası küçük; yalnız başlık içeren ve parçalanmış kaynak chunk'ları inceleyin.",
        "de": "Mehr als ein Viertel der Chunks ist zu klein; reine Überschriften und fragmentierte Quell-Chunks prüfen.",
        "fr": "Plus d’un quart des fragments sont trop petits ; examinez les titres seuls et les fragments source morcelés.",
        "zh-CN": "超过四分之一的分块过小；检查仅标题及碎片化的上游分块。",
        "it": "Oltre un quarto dei chunk è troppo piccolo; esamina titoli isolati e chunk sorgente frammentati.",
    },
    "Tables were observed; compare this profile with table_analytics in a retrieval benchmark.": {
        "tr": "Tablolar gözlendi; getirme karşılaştırmalı testinde bu profili tablo analizi profiliyle karşılaştırın.",
        "de": "Tabellen wurden erkannt; dieses Profil im Abruf-Benchmark mit dem Tabellenanalyseprofil vergleichen.",
        "fr": "Des tableaux ont été détectés ; comparez ce profil au profil d’analyse de tableaux dans un benchmark.",
        "zh-CN": "检测到表格；请在检索基准中与表格分析配置比较。",
        "it": "Sono state rilevate tabelle; confronta il profilo con quello di analisi delle tabelle in un benchmark.",
    },
    "Code blocks were observed; compare this profile with code_assistant and preserve structural boundaries.": {
        "tr": "Kod blokları gözlendi; bu profili kod asistanı profiliyle karşılaştırın ve yapısal sınırları koruyun.",
        "de": "Codeblöcke wurden erkannt; mit dem Code-Assistentenprofil vergleichen und Strukturgrenzen erhalten.",
        "fr": "Des blocs de code ont été détectés ; comparez au profil d’assistant de code et préservez les limites structurelles.",
        "zh-CN": "检测到代码块；请与代码助手配置比较并保留结构边界。",
        "it": "Sono stati rilevati blocchi di codice; confronta con il profilo di assistente al codice e conserva i confini strutturali.",
    },
    "The profile maximum reaches the declared embedding context limit; leave tokenizer overhead and metadata headroom.": {
        "tr": "Profil üst sınırı belirtilen embedding bağlam sınırına ulaşıyor; tokenizer ve metadata için pay bırakın.",
        "de": "Das Profilmaximum erreicht das Embedding-Kontextlimit; Reserve für Tokenizer und Metadaten lassen.",
        "fr": "Le maximum atteint la limite de contexte d’embedding ; gardez une marge pour le tokenizer et les métadonnées.",
        "zh-CN": "配置上限达到嵌入上下文限制；请为 tokenizer 和元数据留出余量。",
        "it": "Il massimo raggiunge il limite del contesto embedding; lascia margine per tokenizer e metadati.",
    },
    "Use this as an initial candidate, then compare at least one smaller and one larger configuration on representative queries.": {
        "tr": "Bunu ilk aday olarak kullanın; ardından temsili sorgularda en az bir küçük ve bir büyük yapılandırmayla karşılaştırın.",
        "de": "Als ersten Kandidaten verwenden und mit mindestens einer kleineren und größeren Konfiguration vergleichen.",
        "fr": "Utilisez ce candidat initial, puis comparez une configuration plus petite et une plus grande.",
        "zh-CN": "将其作为初始候选，并用代表性查询比较至少一个更小和一个更大的配置。",
        "it": "Usalo come candidato iniziale e confronta almeno una configurazione più piccola e una più grande.",
    },
    "context precision and context recall": {
        "tr": "bağlam kesinliği ve bağlam duyarlılığı",
        "de": "Kontextpräzision und Kontext-Recall",
        "fr": "précision et rappel du contexte",
        "zh-CN": "上下文精确率与召回率",
        "it": "precisione e richiamo del contesto",
    },
    "answer faithfulness": {
        "tr": "yanıt sadakati",
        "de": "Antworttreue",
        "fr": "fidélité de la réponse",
        "zh-CN": "回答忠实度",
        "it": "fedeltà della risposta",
    },
    "answer relevance": {
        "tr": "yanıt ilgisi",
        "de": "Antwortrelevanz",
        "fr": "pertinence de la réponse",
        "zh-CN": "回答相关性",
        "it": "rilevanza della risposta",
    },
    "citation correctness": {
        "tr": "atıf doğruluğu",
        "de": "Zitatkorrektheit",
        "fr": "exactitude des citations",
        "zh-CN": "引用正确性",
        "it": "correttezza delle citazioni",
    },
    "latency and retrieved-token cost": {
        "tr": "gecikme ve getirilen-token maliyeti",
        "de": "Latenz und Kosten der abgerufenen Token",
        "fr": "latence et coût des jetons récupérés",
        "zh-CN": "延迟与检索令牌成本",
        "it": "latenza e costo dei token recuperati",
    },
    "There is no universal best chunk size; workload, document structure, tokenizer, embedding model, and query distribution change the optimum.": {
        "tr": "Evrensel bir en iyi chunk boyutu yoktur; iş yükü, belge yapısı, tokenizer, embedding modeli ve sorgu dağılımı optimumu değiştirir.",
        "de": "Es gibt keine universell beste Chunk-Größe; Arbeitslast, Struktur, Tokenizer, Modell und Abfragen verändern das Optimum.",
        "fr": "Il n’existe pas de taille universelle optimale ; charge, structure, tokenizer, modèle et requêtes changent l’optimum.",
        "zh-CN": "不存在通用最佳分块大小；工作负载、结构、tokenizer、模型和查询分布都会改变最优值。",
        "it": "Non esiste una dimensione universale migliore; carico, struttura, tokenizer, modello e query cambiano l’optimum.",
    },
    "Static source analysis cannot prove retrieval or answer quality without representative queries and relevance labels.": {
        "tr": "Statik kaynak analizi, temsili sorgular ve ilgi etiketleri olmadan getirme veya yanıt kalitesini kanıtlayamaz.",
        "de": "Statische Analyse kann Abruf- oder Antwortqualität ohne repräsentative Abfragen und Relevanzlabels nicht belegen.",
        "fr": "L’analyse statique ne peut prouver la qualité sans requêtes représentatives et étiquettes de pertinence.",
        "zh-CN": "没有代表性查询和相关性标签，静态分析无法证明检索或回答质量。",
        "it": "L’analisi statica non può provare la qualità senza query rappresentative ed etichette di rilevanza.",
    },
    "Token counts are model-independent approximations unless a production tokenizer is supplied.": {
        "tr": "Üretim tokenizer'ı sağlanmadıkça token sayıları modelden bağımsız yaklaşık değerlerdir.",
        "de": "Tokenzahlen sind modellunabhängige Näherungen, solange kein Produktionstokenizer vorliegt.",
        "fr": "Les nombres de jetons sont approximatifs sans tokenizer de production.",
        "zh-CN": "除非提供生产 tokenizer，否则令牌数只是与模型无关的近似值。",
        "it": "I conteggi dei token sono approssimazioni senza un tokenizer di produzione.",
    },
    "Fine-grained factual retrieval usually benefits from smaller evidence units.": {
        "tr": "İnce ayrıntılı olgusal getirme genellikle daha küçük kanıt birimlerinden yararlanır.",
        "de": "Feingranulare Faktensuche profitiert meist von kleineren Evidenzeinheiten.",
        "fr": "La recherche factuelle fine bénéficie généralement d’unités de preuve plus petites.",
        "zh-CN": "细粒度事实检索通常更适合较小的证据单元。",
        "it": "Il recupero fattuale dettagliato beneficia in genere di unità di prova più piccole.",
    },
    "Balanced starting point for mixed question types and ordinary documentation.": {
        "tr": "Karma soru türleri ve olağan dokümantasyon için dengeli bir başlangıç noktasıdır.",
        "de": "Ausgewogener Ausgangspunkt für gemischte Fragetypen und übliche Dokumentation.",
        "fr": "Point de départ équilibré pour des questions variées et une documentation courante.",
        "zh-CN": "适用于混合问题类型和普通文档的均衡起点。",
        "it": "Punto di partenza bilanciato per domande miste e documentazione ordinaria.",
    },
    "Procedures need enough neighboring steps and heading context to remain actionable.": {
        "tr": "Prosedürlerin uygulanabilir kalması için yeterli komşu adım ve başlık bağlamı gerekir.",
        "de": "Verfahren benötigen genügend benachbarte Schritte und Überschriftenkontext.",
        "fr": "Les procédures ont besoin d’étapes voisines et du contexte des titres pour rester applicables.",
        "zh-CN": "流程需要足够的相邻步骤和标题上下文，才能保持可操作性。",
        "it": "Le procedure richiedono passaggi vicini e contesto dei titoli per restare applicabili.",
    },
    "Broad synthesis questions need larger coherent evidence windows.": {
        "tr": "Geniş sentez soruları daha büyük ve tutarlı kanıt pencereleri gerektirir.",
        "de": "Breite Synthesefragen benötigen größere zusammenhängende Evidenzfenster.",
        "fr": "Les questions de synthèse larges exigent des fenêtres de preuve cohérentes plus grandes.",
        "zh-CN": "广泛的综合问题需要更大的连贯证据窗口。",
        "it": "Le domande di sintesi ampie richiedono finestre di prova coerenti più grandi.",
    },
    "Code boundaries and related declarations matter more than arbitrary fixed windows.": {
        "tr": "Kod sınırları ve ilişkili bildirimler, rastgele sabit pencerelerden daha önemlidir.",
        "de": "Codegrenzen und zugehörige Deklarationen sind wichtiger als beliebige feste Fenster.",
        "fr": "Les limites du code et les déclarations liées priment sur des fenêtres fixes arbitraires.",
        "zh-CN": "代码边界和相关声明比任意固定窗口更重要。",
        "it": "I confini del codice e le dichiarazioni correlate contano più di finestre fisse arbitrarie.",
    },
    "Tables should remain intact; duplicated row overlap can distort numeric retrieval.": {
        "tr": "Tablolar bütün kalmalıdır; yinelenen satır bindirmesi sayısal getirmeyi bozabilir.",
        "de": "Tabellen sollten intakt bleiben; überlappende Zeilen können numerischen Abruf verzerren.",
        "fr": "Les tableaux doivent rester intacts ; le chevauchement des lignes peut fausser la recherche numérique.",
        "zh-CN": "表格应保持完整；重复行重叠可能扭曲数值检索。",
        "it": "Le tabelle devono restare integre; righe sovrapposte possono distorcere il recupero numerico.",
    },
}
for _source_text, _localized_values in _RAG_NARRATIVE_TRANSLATIONS.items():
    for _rag_locale, _localized_text in _localized_values.items():
        _TRANSLATIONS[_rag_locale][_source_text] = _localized_text

_DUPLICATE_TRANSLATIONS: dict[str, dict[str, str]] = {
    "tr": {
        "Duplicate comparisons": "Yinelenen içerik karşılaştırmaları",
        "Other findings": "Diğer bulgular",
        "Duplicate Comparison": "Yinelenen Karşılaştırma",
        "Compared locations": "Karşılaştırılan konumlar",
        "Reference occurrence": "Referans tekrar",
        "Matching occurrence": "Eşleşen tekrar",
        "Exact match after normalization": "Normalleştirme sonrası tam eşleşme",
        "Near match requiring manual review": "Elle inceleme gerektiren yakın eşleşme",
        "Same-document repetition": "Aynı belge içindeki tekrar",
        "Similarity": "Benzerlik",
        "Estimated redundant tokens": "Tahmini yinelenen token",
        "Compared excerpt": "Karşılaştırılan metin",
        "Shared phrases": "Ortak ifadeler",
        "Review both locations together. A match is not an automatic deletion decision.": "Her iki konumu birlikte inceleyin. Bir eşleşme otomatik silme kararı değildir.",
        "Comparison details are unavailable in this older report. Run the scan again.": "Bu eski raporda karşılaştırma ayrıntıları yok. Taramayı yeniden çalıştırın.",
        "Near-duplicate content group": "Yakın yinelenen içerik grubu",
        "Near-identical content may waste retrieval capacity or over-weight one statement.": "Neredeyse aynı içerik, getirme kapasitesini tüketebilir veya bir ifadeye gereğinden fazla ağırlık verebilir.",
        "Review the group manually; similarity is not proof that an item should be deleted.": "İki tarafı birlikte elle inceleyin; benzerlik tek başına bir içeriğin silinmesi gerektiğini kanıtlamaz.",
    },
    "de": {
        "Duplicate comparisons": "Duplikatvergleiche",
        "Other findings": "Weitere Befunde",
        "Duplicate Comparison": "Duplikatvergleich",
        "Compared locations": "Verglichene Fundstellen",
        "Reference occurrence": "Referenzvorkommen",
        "Matching occurrence": "Übereinstimmendes Vorkommen",
        "Exact match after normalization": "Exakte Übereinstimmung nach Normalisierung",
        "Near match requiring manual review": "Ähnliche Fundstelle zur manuellen Prüfung",
        "Same-document repetition": "Wiederholung im selben Dokument",
        "Similarity": "Ähnlichkeit",
        "Estimated redundant tokens": "Geschätzte redundante Token",
        "Compared excerpt": "Verglichener Auszug",
        "Shared phrases": "Gemeinsame Formulierungen",
        "Review both locations together. A match is not an automatic deletion decision.": "Beide Fundstellen gemeinsam prüfen. Eine Übereinstimmung ist keine automatische Löschentscheidung.",
        "Comparison details are unavailable in this older report. Run the scan again.": "In diesem älteren Bericht fehlen Vergleichsdetails. Scan erneut ausführen.",
        "Near-duplicate content group": "Gruppe nahezu identischer Inhalte",
        "Near-identical content may waste retrieval capacity or over-weight one statement.": "Nahezu identischer Inhalt kann Abrufkapazität verschwenden oder eine Aussage übergewichten.",
        "Review the group manually; similarity is not proof that an item should be deleted.": "Beide Fundstellen manuell prüfen; Ähnlichkeit ist kein Beweis für eine Löschung.",
    },
    "fr": {
        "Duplicate comparisons": "Comparaisons des doublons",
        "Other findings": "Autres constats",
        "Duplicate Comparison": "Comparaison doublons",
        "Compared locations": "Emplacements comparés",
        "Reference occurrence": "Occurrence de référence",
        "Matching occurrence": "Occurrence correspondante",
        "Exact match after normalization": "Correspondance exacte après normalisation",
        "Near match requiring manual review": "Correspondance proche à vérifier manuellement",
        "Same-document repetition": "Répétition dans le même document",
        "Similarity": "Similarité",
        "Estimated redundant tokens": "Jetons redondants estimés",
        "Compared excerpt": "Extrait comparé",
        "Shared phrases": "Expressions communes",
        "Review both locations together. A match is not an automatic deletion decision.": "Examinez les deux emplacements ensemble. Une correspondance n’entraîne pas une suppression automatique.",
        "Comparison details are unavailable in this older report. Run the scan again.": "Les détails de comparaison manquent dans cet ancien rapport. Relancez l’analyse.",
        "Near-duplicate content group": "Groupe de contenus quasi dupliqués",
        "Near-identical content may waste retrieval capacity or over-weight one statement.": "Un contenu quasi identique peut gaspiller la capacité de recherche ou surpondérer une affirmation.",
        "Review the group manually; similarity is not proof that an item should be deleted.": "Examinez les deux côtés manuellement ; la similarité ne prouve pas qu’un élément doit être supprimé.",
    },
    "zh-CN": {
        "Duplicate comparisons": "重复内容对比",
        "Other findings": "其他发现",
        "Duplicate Comparison": "重复对比",
        "Compared locations": "对比位置",
        "Reference occurrence": "参考位置",
        "Matching occurrence": "匹配位置",
        "Exact match after normalization": "规范化后完全匹配",
        "Near match requiring manual review": "需要人工审核的近似匹配",
        "Same-document repetition": "同一文档内重复",
        "Similarity": "相似度",
        "Estimated redundant tokens": "估算冗余令牌",
        "Compared excerpt": "对比文本",
        "Shared phrases": "共同短语",
        "Review both locations together. A match is not an automatic deletion decision.": "请同时检查两个位置。匹配并不等于自动删除。",
        "Comparison details are unavailable in this older report. Run the scan again.": "此旧报告没有对比详情，请重新运行扫描。",
        "Near-duplicate content group": "近似重复内容组",
        "Near-identical content may waste retrieval capacity or over-weight one statement.": "近似重复内容可能浪费检索容量或使某一陈述权重过高。",
        "Review the group manually; similarity is not proof that an item should be deleted.": "请人工对比两侧；相似度本身不能证明应删除某项。",
    },
    "it": {
        "Duplicate comparisons": "Confronti dei duplicati",
        "Other findings": "Altri risultati",
        "Duplicate Comparison": "Confronto duplicati",
        "Compared locations": "Posizioni confrontate",
        "Reference occurrence": "Occorrenza di riferimento",
        "Matching occurrence": "Occorrenza corrispondente",
        "Exact match after normalization": "Corrispondenza esatta dopo la normalizzazione",
        "Near match requiring manual review": "Corrispondenza simile da verificare manualmente",
        "Same-document repetition": "Ripetizione nello stesso documento",
        "Similarity": "Somiglianza",
        "Estimated redundant tokens": "Token ridondanti stimati",
        "Compared excerpt": "Estratto confrontato",
        "Shared phrases": "Espressioni comuni",
        "Review both locations together. A match is not an automatic deletion decision.": "Esamina insieme entrambe le posizioni. Una corrispondenza non implica l’eliminazione automatica.",
        "Comparison details are unavailable in this older report. Run the scan again.": "I dettagli non sono disponibili in questo vecchio rapporto. Esegui di nuovo la scansione.",
        "Near-duplicate content group": "Gruppo di contenuti quasi duplicati",
        "Near-identical content may waste retrieval capacity or over-weight one statement.": "Contenuti quasi identici possono sprecare capacità di recupero o sovrappesare un’affermazione.",
        "Review the group manually; similarity is not proof that an item should be deleted.": "Confronta manualmente entrambi i lati; la somiglianza non dimostra che un elemento vada eliminato.",
    },
}
for _duplicate_locale, _duplicate_catalog in _DUPLICATE_TRANSLATIONS.items():
    _TRANSLATIONS[_duplicate_locale].update(_duplicate_catalog)

_DUPLICATE_DETAIL_TRANSLATIONS: dict[str, dict[str, str]] = {
    "tr": {
        "Matched content": "Eşleşen içerik",
        "Affected chunks": "Etkilenen parça sayısı",
        "Group count": "Grup sayısı",
        "Within-document near-duplicate chunks": "Belge içinde yakın yinelenen parçalar",
        "Within-document near match": "Belge içindeki yakın eşleşme",
        "{count} more affected chunks omitted; use HTML or Excel for the complete list.": (
            "{count} etkilenen parça daha PDF özetinden çıkarıldı; tam liste için HTML "
            "veya Excel kullanın."
        ),
    },
    "de": {
        "Matched content": "Übereinstimmender Inhalt",
        "Affected chunks": "Betroffene Chunks",
        "Group count": "Gruppenanzahl",
        "Within-document near-duplicate chunks": "Nahezu identische Chunks im Dokument",
        "Within-document near match": "Ähnliche Stelle im selben Dokument",
        "{count} more affected chunks omitted; use HTML or Excel for the complete list.": (
            "{count} weitere betroffene Chunks wurden ausgelassen; vollständige Liste in "
            "HTML oder Excel."
        ),
    },
    "fr": {
        "Matched content": "Contenu correspondant",
        "Affected chunks": "Fragments concernés",
        "Group count": "Nombre de groupes",
        "Within-document near-duplicate chunks": "Fragments quasi dupliqués dans le document",
        "Within-document near match": "Correspondance proche dans le document",
        "{count} more affected chunks omitted; use HTML or Excel for the complete list.": (
            "{count} fragments concernés supplémentaires ont été omis ; liste complète en "
            "HTML ou Excel."
        ),
    },
    "zh-CN": {
        "Matched content": "匹配内容",
        "Affected chunks": "受影响分块数",
        "Group count": "组数",
        "Within-document near-duplicate chunks": "文档内近似重复分块",
        "Within-document near match": "文档内近似匹配",
        "{count} more affected chunks omitted; use HTML or Excel for the complete list.": (
            "另有 {count} 个受影响分块未在 PDF 摘要中显示；完整列表请使用 HTML 或 Excel。"
        ),
    },
    "it": {
        "Matched content": "Contenuto corrispondente",
        "Affected chunks": "Chunk interessati",
        "Group count": "Numero di gruppi",
        "Within-document near-duplicate chunks": "Chunk quasi duplicati nel documento",
        "Within-document near match": "Corrispondenza vicina nello stesso documento",
        "{count} more affected chunks omitted; use HTML or Excel for the complete list.": (
            "{count} ulteriori chunk interessati sono stati omessi; elenco completo in HTML "
            "o Excel."
        ),
    },
}
for _detail_locale, _detail_catalog in _DUPLICATE_DETAIL_TRANSLATIONS.items():
    _TRANSLATIONS[_detail_locale].update(_detail_catalog)

_QUALITY_RECOMMENDATION_TRANSLATIONS: dict[str, dict[str, str]] = {
    "tr": {
        "Rechunk with safer structural boundaries and review the affected source.": "Daha güvenli yapısal sınırlarla yeniden chunk'layın ve etkilenen kaynağı inceleyin.",
        "Consider merging with a related adjacent chunk while preserving headings.": "Başlıkları koruyarak ilişkili komşu chunk ile birleştirmeyi değerlendirin.",
        "Reparse the source and inspect extraction/mapping warnings.": "Kaynağı yeniden ayrıştırın; çıkarım ve kaynak eşleme uyarılarını inceleyin.",
        "Review boilerplate policy before indexing; do not remove content automatically.": "İndekslemeden önce şablon metin politikasını inceleyin; içeriği otomatik kaldırmayın.",
        "Review the chunk and adjust deterministic chunking configuration if appropriate.": "Chunk'ı inceleyin ve uygunsa deterministik chunk ayarlarını düzeltin.",
    },
    "de": {
        "Rechunk with safer structural boundaries and review the affected source.": "Mit sichereren Strukturgrenzen neu aufteilen und die betroffene Quelle prüfen.",
        "Consider merging with a related adjacent chunk while preserving headings.": "Zusammenführung mit einem benachbarten Chunk unter Erhalt der Überschriften prüfen.",
        "Reparse the source and inspect extraction/mapping warnings.": "Quelle erneut parsen und Extraktions- sowie Zuordnungswarnungen prüfen.",
        "Review boilerplate policy before indexing; do not remove content automatically.": "Standardtextregeln vor der Indexierung prüfen; Inhalte nicht automatisch entfernen.",
        "Review the chunk and adjust deterministic chunking configuration if appropriate.": "Chunk prüfen und die deterministische Konfiguration bei Bedarf anpassen.",
    },
    "fr": {
        "Rechunk with safer structural boundaries and review the affected source.": "Redécoupez avec des limites structurelles plus sûres et examinez la source concernée.",
        "Consider merging with a related adjacent chunk while preserving headings.": "Envisagez la fusion avec un fragment voisin lié tout en conservant les titres.",
        "Reparse the source and inspect extraction/mapping warnings.": "Analysez de nouveau la source et examinez les avertissements d’extraction et de mappage.",
        "Review boilerplate policy before indexing; do not remove content automatically.": "Vérifiez la politique de texte standard avant l’indexation ; ne supprimez rien automatiquement.",
        "Review the chunk and adjust deterministic chunking configuration if appropriate.": "Examinez le fragment et ajustez la configuration déterministe si nécessaire.",
    },
    "zh-CN": {
        "Rechunk with safer structural boundaries and review the affected source.": "使用更安全的结构边界重新分块，并检查受影响的来源。",
        "Consider merging with a related adjacent chunk while preserving headings.": "在保留标题的前提下，考虑与相关的相邻分块合并。",
        "Reparse the source and inspect extraction/mapping warnings.": "重新解析来源，并检查提取和映射警告。",
        "Review boilerplate policy before indexing; do not remove content automatically.": "索引前检查模板文本策略；不要自动删除内容。",
        "Review the chunk and adjust deterministic chunking configuration if appropriate.": "检查分块，并按需调整确定性分块配置。",
    },
    "it": {
        "Rechunk with safer structural boundaries and review the affected source.": "Suddividi di nuovo con confini strutturali più sicuri e verifica la fonte interessata.",
        "Consider merging with a related adjacent chunk while preserving headings.": "Valuta l’unione con un chunk adiacente correlato conservando i titoli.",
        "Reparse the source and inspect extraction/mapping warnings.": "Analizza nuovamente la fonte e verifica gli avvisi di estrazione e mappatura.",
        "Review boilerplate policy before indexing; do not remove content automatically.": "Verifica la policy del testo standard prima dell’indicizzazione; non rimuovere automaticamente.",
        "Review the chunk and adjust deterministic chunking configuration if appropriate.": "Esamina il chunk e modifica la configurazione deterministica se necessario.",
    },
}
for _quality_locale, _quality_catalog in _QUALITY_RECOMMENDATION_TRANSLATIONS.items():
    _TRANSLATIONS[_quality_locale].update(_quality_catalog)


def _locale(value: str) -> str:
    return value if value in _LOCALES else "en"


def _translator(locale: str):  # type: ignore[no-untyped-def]
    catalog = _TRANSLATIONS.get(_locale(locale), {})
    return lambda value: catalog.get(value, value)


def _display(value: Any, *, fallback: str) -> str:
    if value is None or value == "":
        return fallback
    if isinstance(value, float):
        return f"{value:.1f}"
    if isinstance(value, datetime):
        return value.astimezone().strftime("%Y-%m-%d %H:%M %Z")
    return str(value)


def _safe_filename(value: str) -> str:
    normalized = re.sub(r"[^A-Za-z0-9_-]+", "-", value).strip("-_")
    return (normalized[:80] or "report").lower()


def report_export_filename(report: ReportDocument, history_id: str, extension: str) -> str:
    source = _safe_filename(str(report.scan.get("source_name") or "report"))
    identifier = _safe_filename(history_id[:12])
    return f"ragscanner-{source}-{identifier}.{extension}"


def export_report(
    report: ReportDocument,
    export_format: ReportExportFormat,
    *,
    locale: str = "en",
) -> ReportExport:
    if export_format == "html":
        return ReportExport(_render_html(report, locale).encode("utf-8"), "text/html", "html")
    if export_format == "xlsx":
        return ReportExport(
            _render_xlsx(report, locale),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xlsx",
        )
    if export_format == "pdf":
        return ReportExport(_render_pdf(report, locale), "application/pdf", "pdf")
    raise ValueError("unsupported report export format")


def _score_band(value: float | None) -> str:
    if value is None:
        return "unassessed"
    if value < 55:
        return "critical"
    if value < 70:
        return "poor"
    if value < 85:
        return "warning"
    return "healthy"


def _finding_location(finding: ReportFinding, t) -> str:  # type: ignore[no-untyped-def]
    parts = [finding.source or t("Not assessed")]
    if finding.page is not None:
        parts.append(f"{t('Page')} {finding.page}")
    if finding.line_start is not None:
        line = str(finding.line_start)
        if finding.line_end is not None and finding.line_end != finding.line_start:
            line += f"-{finding.line_end}"
        parts.append(f"{t('Line')} {line}")
    return " · ".join(parts)


def _duplicate_kind(category: str) -> str:
    if category == "repeated_chunk_within_document":
        return "Same-document repetition"
    if category == "within_document_near_duplicates":
        return "Within-document near match"
    if category.startswith("near_duplicate"):
        return "Near match requiring manual review"
    return "Exact match after normalization"


def _duplicate_member_location(member: Any, t) -> str:  # type: ignore[no-untyped-def]
    parts = [member.source or t("Not assessed")]
    if member.page is not None:
        parts.append(f"{t('Page')} {member.page}")
    if member.line_start is not None:
        line = str(member.line_start)
        if member.line_end is not None and member.line_end != member.line_start:
            line += f"–{member.line_end}"
        parts.append(f"{t('Line')} {line}")
    return " · ".join(parts)


_COVERAGE_AREA_TEXT = {
    "chunk_quality": "Chunk quality",
    "cross_document_exact_duplicates": "Cross-document exact duplicates",
    "cross_document_freshness": "Cross-document freshness",
    "cross_document_near_duplicates": "Cross-document near duplicates",
    "static_security": "Static security",
    "version_conflict": "Version conflicts",
    "within_document_near_duplicates": "Within-document near duplicates",
    "within_document_repeated_chunks": "Within-document repeated chunks",
}


def _coverage_area(value: str, t) -> str:  # type: ignore[no-untyped-def]
    return cast(str, t(_COVERAGE_AREA_TEXT.get(value, value.replace("_", " ").title())))


def _ai_root_cause_text(item: Any, t) -> str:  # type: ignore[no-untyped-def]
    rules = ", ".join(item.finding_rules)
    files = ", ".join(item.example_files)
    return (
        f"{item.pattern} · {item.label} · {t(item.confidence)}\n"
        f"{t('Rule')}: {rules}\n{t('File')}: {files}\n{item.explanation}"
    )


def _ai_action_text(item: Any, t) -> str:  # type: ignore[no-untyped-def]
    addressed = ", ".join(item.addresses) or "—"
    return (
        f"{item.order}. {item.action}\n"
        f"{t('Target')}: {t(item.target)} · {t('Effort')}: {t(item.effort)}\n"
        f"{t('Addresses')}: {addressed}\n"
        f"{t('Expected effect')}: {item.expected_effect}"
    )


def _ai_question_text(item: Any, t) -> str:  # type: ignore[no-untyped-def]
    return f"{item.question}\n{t('Informs')}: {item.informs}"


def _render_html(report: ReportDocument, locale: str) -> str:
    locale = _locale(locale)
    t = _translator(locale)

    def esc(value: Any) -> str:
        return html.escape(_display(value, fallback=t("Not assessed")), quote=True)

    score_labels = {
        "overall": t("Overall score"),
        "security": t("Security score"),
        "knowledge_quality": t("Content quality"),
        "efficiency": t("Efficiency"),
    }
    scores = "".join(
        f'<article class="score {_score_band(report.scores.get(key))}"><span>{esc(label)}</span>'
        f"<strong>{esc(report.scores.get(key))}</strong></article>"
        for key, label in score_labels.items()
    )
    severities = "".join(
        f"<div><span>{esc(t(name.title()))}</span><strong>{report.severity_summary.get(name, 0)}</strong></div>"
        for name in ("critical", "high", "medium", "low", "info")
    )
    actions = (
        {item.finding_id: item for item in report.ai_analysis.finding_actions}
        if report.ai_analysis
        else {}
    )
    findings = []
    for finding in report.findings:
        if finding.metadata.get("group_id"):
            continue
        action = actions.get(finding.id)
        title = localize_rule_field(finding.rule_id, locale, "title", finding.title)
        impact = localize_rule_field(finding.rule_id, locale, "impact", finding.impact)
        remediation = (
            action.remediation
            if action
            else localize_rule_field(
                finding.rule_id,
                locale,
                "recommendation",
                finding.recommendation,
            )
        )
        verification = (
            "<ol>"
            + "".join(f"<li>{esc(step)}</li>" for step in action.verification_steps)
            + "</ol>"
            if action and action.verification_steps
            else ""
        )
        highlight = (
            f"<p><strong>{esc(t('Problematic text'))}</strong><br><mark>{esc(finding.evidence_highlight)}</mark></p>"
            if finding.evidence_highlight
            else ""
        )
        findings.append(
            f'<details class="finding" open><summary><span class="badge {esc(finding.severity.value)}">{esc(t(finding.severity.value.title()))}</span> '
            f"{esc(title)} <code>{esc(finding.rule_id)}</code></summary>"
            f"<p><strong>{esc(t('Source location'))}:</strong> {esc(_finding_location(finding, t))}</p>"
            f'<div class="columns"><section><h3>{esc(t("Impact"))}</h3><p>{esc(impact)}</p></section>'
            f"<section><h3>{esc(t('Evidence'))}</h3>{highlight}<blockquote>{esc(finding.evidence)}</blockquote></section>"
            f"<section><h3>{esc(t('AI-assisted fix') if action else t('Recommendation'))}</h3><p>{esc(remediation)}</p>{verification}</section></div></details>"
        )
    ai_section = ""
    if report.ai_analysis:
        ai = report.ai_analysis
        roots = "".join(
            f"<li><strong>{esc(item.pattern)} · {esc(item.label)}</strong> "
            f"<span>({esc(t(item.confidence))})</span><br>"
            f"{esc(t('Rule'))}: {esc(', '.join(item.finding_rules))}<br>"
            f"{esc(t('File'))}: {esc(', '.join(item.example_files))}<br>"
            f"{esc(item.explanation)}</li>"
            for item in ai.root_causes
        )
        priority = "".join(
            f"<li><strong>{esc(item.action)}</strong><br>"
            f"{esc(t('Target'))}: {esc(t(item.target))} · "
            f"{esc(t('Effort'))}: {esc(t(item.effort))}<br>"
            f"{esc(t('Addresses'))}: {esc(', '.join(item.addresses) or '—')}<br>"
            f"{esc(t('Expected effect'))}: {esc(item.expected_effect)}</li>"
            for item in sorted(ai.priority_actions, key=lambda value: value.order)
        )
        questions = "".join(
            f"<li>{esc(item.question)}<br><small>{esc(t('Informs'))}: "
            f"{esc(item.informs)}</small></li>"
            for item in ai.review_questions
        )
        caveat = (
            f"<p><em><strong>{esc(t('Coverage caveat'))}:</strong> "
            f"{esc(ai.coverage_caveat)}</em></p>"
            if ai.coverage_caveat
            else ""
        )
        ai_section = (
            f"<section><h2>{esc(t('AI-assisted analysis'))}</h2><p>{esc(ai.ai_analysis)}</p>"
            f"<h3>{esc(t('Score commentary'))}</h3><p>{esc(ai.score_commentary)}</p>{caveat}"
            f"<h3>{esc(t('Root cause analysis'))}</h3><ol>{roots}</ol>"
            f'<div class="columns"><div><h3>{esc(t("Priority actions"))}</h3><ol>{priority}</ol></div>'
            f"<div><h3>{esc(t('Questions for review'))}</h3><ul>{questions}</ul></div></div>"
            f'<p class="muted">{esc(ai.provider)} · {esc(ai.model)} · {esc(t(ai.disclaimer))}</p></section>'
        )
    elif report.ai_analysis_error:
        ai_section = (
            f"<section><h2>{esc(t('AI analysis unavailable'))}</h2>"
            f"<p><code>{esc(report.ai_analysis_error_code)}</code> {esc(report.ai_analysis_error)}</p></section>"
        )
    rag_section = ""
    if report.rag_configuration_advice:
        rag = report.rag_configuration_advice
        recommended = rag.recommended
        rag_why = t(str(recommended.get("why") or "Not assessed"))
        rag_section = (
            f"<section><h2>{esc(t('RAG configuration advice'))}</h2>"
            f"<p><strong>{esc(t('Workload profile'))}:</strong> {esc(t(rag.profile.value))}</p>"
            f'<div class="columns"><div><h3>{esc(t("Recommended chunk range"))}</h3>'
            f"<p>{esc(recommended.get('minimum_tokens'))}–{esc(recommended.get('maximum_tokens'))} "
            f"{esc(t('tokens'))}; {esc(t('target'))} {esc(recommended.get('target_tokens'))}</p></div>"
            f"<div><h3>{esc(t('Overlap'))}</h3><p>{esc(recommended.get('overlap_tokens'))} "
            f"{esc(t('tokens'))}</p></div><div><h3>{esc(t('Retrieval top-k'))}</h3>"
            f"<p>{esc(recommended.get('retrieval_top_k'))}</p></div></div>"
            f"<h3>{esc(t('Why this range?'))}</h3><p>{esc(rag_why)}</p>"
            f"<h3>{esc(t('Actions'))}</h3><ul>{''.join(f'<li>{esc(t(item))}</li>' for item in rag.actions)}</ul>"
            f"<h3>{esc(t('Validation metrics'))}</h3><ul>{''.join(f'<li>{esc(t(item))}</li>' for item in rag.validation_metrics)}</ul>"
            f'<p class="muted">{esc(t("Starting point only; validate with representative queries before production use."))}</p></section>'
        )
    duplicate_section = ""
    if report.duplicate_groups:
        group_cards: list[str] = []
        for group in report.duplicate_groups:
            occurrences = []
            for member in group.members:
                role = "Reference occurrence" if member.canonical else "Matching occurrence"
                excerpt = (
                    f"<mark>{esc(member.evidence_excerpt)}</mark>"
                    if member.evidence_excerpt
                    else f'<span class="muted">{esc(t("Not assessed"))}</span>'
                )
                occurrences.append(
                    f'<article class="occurrence"><h3>{esc(t(role))}</h3>'
                    f"<p><strong>{esc(_duplicate_member_location(member, t))}</strong></p>"
                    f'<p class="muted">{esc(t("Compared excerpt"))}</p><blockquote>{excerpt}</blockquote></article>'
                )
            comparison = (
                '<div class="occurrence-grid">' + "".join(occurrences) + "</div>"
                if occurrences
                else f"<p>{esc(t('Comparison details are unavailable in this older report. Run the scan again.'))}</p>"
            )
            shared = (
                f"<p><strong>{esc(t('Shared phrases'))}:</strong> "
                + " ".join(f"<mark>{esc(value)}</mark>" for value in group.shared_phrases)
                + "</p>"
                if group.shared_phrases
                else ""
            )
            matched = (
                f"<p><strong>{esc(t('Matched content'))}:</strong> "
                f"<mark>{esc(group.matched_content)}</mark></p>"
                if group.matched_content
                else ""
            )
            group_cards.append(
                f'<details class="duplicate-group" open><summary>{esc(t(_duplicate_kind(group.category)))}'
                f" · {esc(t('Similarity'))} {group.similarity * 100:.1f}%"
                f" · {esc(t('Affected chunks'))}: {group.affected_chunks}"
                f" · {esc(t('Group count'))}: {group.group_count}</summary>"
                f"{matched}"
                f'<p class="muted">{esc(t("Estimated redundant tokens"))}: {group.estimated_redundant_tokens}</p>'
                f"{shared}{comparison}</details>"
            )
        duplicate_section = (
            f"<section><h2>{esc(t('Duplicate comparisons'))}</h2>"
            f'<p class="muted">{esc(t("Review both locations together. A match is not an automatic deletion decision."))}</p>'
            + "".join(group_cards)
            + "</section>"
        )
    coverage_rows = "".join(
        f"<tr><td>{esc(_coverage_area(area, t))}</td><td>{esc(t(str(value.get('status'))))}</td>"
        f"<td>{esc(localize_coverage_reason(str(value.get('reason') or ''), locale))}</td></tr>"
        for area, value in sorted(report.assessment_coverage.items())
    )
    ingestion_rows = "".join(
        f"<tr><td>{esc(item.path)}</td><td>{esc(item.stage)}</td><td>{esc(item.message)}</td><td>{esc(item.remediation)}</td></tr>"
        for item in report.ingestion_issues
    )
    return f'''<!doctype html><html lang="{html.escape(locale, quote=True)}"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><meta http-equiv="Content-Security-Policy" content="default-src 'none'; style-src 'unsafe-inline'; img-src data:; font-src 'none'; script-src 'none'; connect-src 'none'; base-uri 'none'; form-action 'none'"><title>{esc(t("RAGScanner report"))}</title><style>
:root{{--ink:#10233d;--muted:#5f6d7d;--line:#dce4ea;--accent:#078c91;--panel:#fff;--bg:#f3f7f9}}*{{box-sizing:border-box}}body{{margin:0;background:var(--bg);color:var(--ink);font:15px/1.55 system-ui,sans-serif}}header{{padding:32px;background:linear-gradient(120deg,#061a35,#075b70);color:#fff}}header>div,main,footer{{max-width:1180px;margin:auto}}main,footer{{padding:24px}}section,.finding{{background:var(--panel);border:1px solid var(--line);border-radius:14px;padding:20px;margin:16px 0}}h1,h2,h3{{line-height:1.2}}.meta,.score-grid,.severity,.columns{{display:grid;gap:12px}}.meta{{grid-template-columns:repeat(3,1fr)}}.score-grid{{grid-template-columns:repeat(4,1fr)}}.severity{{grid-template-columns:repeat(5,1fr)}}.columns{{grid-template-columns:repeat(3,1fr)}}.score,.severity>div{{border:1px solid var(--line);border-radius:10px;padding:14px}}.score strong,.severity strong{{display:block;font-size:24px}}.healthy{{border-top:5px solid #15935a}}.warning{{border-top:5px solid #e2b100;background:#fffbeb}}.poor{{border-top:5px solid #e17016;background:#fff4e8}}.critical{{border-top:5px solid #d23845;background:#fff0f1}}.unassessed{{border-top:5px solid #84909c}}.badge{{border:1px solid currentColor;border-radius:99px;padding:2px 8px;font-weight:700}}summary{{cursor:pointer;font-weight:700}}blockquote,mark{{overflow-wrap:anywhere}}blockquote{{margin:8px 0;padding:12px;border-left:4px solid var(--accent);background:#f6fafb}}mark{{background:#fff1a8;padding:2px}}.duplicate-group{{border:1px solid var(--line);border-radius:10px;padding:14px;margin-top:12px}}.occurrence-grid{{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:12px}}.occurrence{{border:1px solid var(--line);border-radius:8px;padding:14px;min-width:0}}.occurrence h3{{margin-top:0}}table{{width:100%;border-collapse:collapse}}th,td{{padding:9px;border-bottom:1px solid var(--line);text-align:left;vertical-align:top}}.muted{{color:var(--muted)}}code{{overflow-wrap:anywhere}}@media(max-width:760px){{.meta,.score-grid,.severity,.columns,.occurrence-grid{{grid-template-columns:1fr}}table{{display:block;overflow:auto}}main,footer{{padding:14px}}}}@media print{{body{{background:#fff}}header{{background:#fff;color:#000;border-bottom:2px solid #000}}section,.finding{{break-inside:avoid}}}}
</style></head><body><header><div><p>RAGScanner</p><h1>{esc(t("RAGScanner report"))}</h1><p>{esc(report.scan.get("id"))}</p></div></header><main><section><h2>{esc(t("Executive summary"))}</h2><div class="meta"><p><strong>{esc(t("Source"))}</strong><br>{esc(report.scan.get("source_name"))}</p><p><strong>{esc(t("Status"))}</strong><br>{esc(t(str(report.scan.get("status"))))}</p><p><strong>{esc(t("Generated"))}</strong><br>{esc(report.generated_at)}</p></div><div class="score-grid">{scores}</div><div class="severity">{severities}</div></section>{rag_section}{ai_section}{duplicate_section}<section><h2>{esc(t("Other findings") if report.duplicate_groups else t("Findings"))}</h2>{"".join(findings) or f"<p>{esc(t('No findings recorded.'))}</p>"}</section><section><h2>{esc(t("Coverage"))}</h2><table><thead><tr><th>{esc(t("Area"))}</th><th>{esc(t("Status"))}</th><th>{esc(t("Reason"))}</th></tr></thead><tbody>{coverage_rows}</tbody></table></section><section><h2>{esc(t("Ingestion issues"))}</h2>{f"<table><thead><tr><th>{esc(t('Path'))}</th><th>{esc(t('Stage'))}</th><th>{esc(t('Issue'))}</th><th>{esc(t('Remediation'))}</th></tr></thead><tbody>{ingestion_rows}</tbody></table>" if ingestion_rows else f"<p>{esc(t('No ingestion issues recorded.'))}</p>"}</section></main><footer><p>{esc(t("Generated locally by RAGScanner. No external assets or network requests."))}</p></footer></body></html>'''


def _cell(value: Any) -> str | int | float:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)):
        return value
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", str(value))[:_MAX_CELL_LENGTH]
    if text.startswith(("=", "+", "-", "@")):
        text = "'" + text
    return text


def _style_sheet(sheet, *, freeze: str | None = None, auto_filter: bool = False) -> None:  # type: ignore[no-untyped-def]
    header_fill = PatternFill("solid", fgColor="075B70")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    if freeze:
        sheet.freeze_panes = freeze
    if auto_filter and sheet.max_row > 1:
        sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(sheet.columns, start=1):
        width = min(60, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        sheet.column_dimensions[get_column_letter(index)].width = width


def _append_key_value(sheet, key: str, value: Any) -> None:  # type: ignore[no-untyped-def]
    sheet.append([_cell(key), _cell(value)])


def _render_xlsx(report: ReportDocument, locale: str) -> bytes:
    t = _translator(locale)
    workbook = Workbook()
    summary = cast(Worksheet, workbook.active)
    summary.title = t("Summary")
    summary.append([t("Report details"), ""])
    _append_key_value(summary, t("Source"), report.scan.get("source_name"))
    _append_key_value(summary, t("Status"), t(str(report.scan.get("status"))))
    _append_key_value(summary, t("Generated"), _display(report.generated_at, fallback=""))
    _append_key_value(summary, t("Overall score"), report.scores.get("overall"))
    _append_key_value(summary, t("Security score"), report.scores.get("security"))
    _append_key_value(summary, t("Content quality"), report.scores.get("knowledge_quality"))
    _append_key_value(summary, t("Efficiency"), report.scores.get("efficiency"))
    _append_key_value(summary, t("Files discovered"), report.processing.files_discovered)
    _append_key_value(summary, t("Files processed"), report.processing.files_scanned)
    _append_key_value(summary, t("Files skipped"), report.processing.files_skipped)
    _append_key_value(summary, t("Findings"), len(report.findings))
    for severity in ("critical", "high", "medium", "low", "info"):
        _append_key_value(summary, t(severity.title()), report.severity_summary.get(severity, 0))
    score = report.scores.get("overall")
    if score is not None:
        fill = {"healthy": "DDF3E5", "warning": "FFF1B8", "poor": "FFD8B5", "critical": "FFC7CE"}[
            _score_band(score)
        ]
        summary["A5"].fill = PatternFill("solid", fgColor=fill)
        summary["B5"].fill = PatternFill("solid", fgColor=fill)
    _style_sheet(summary)

    finding_sheet = workbook.create_sheet(t("Findings"))
    finding_sheet.append(
        [
            t("Severity"),
            t("Rule"),
            t("Title"),
            t("Category"),
            t("Confidence"),
            t("File"),
            t("Page"),
            t("Line"),
            t("Problematic text"),
            t("Evidence"),
            t("Impact"),
            t("Recommendation"),
            t("AI-assisted fix"),
            t("Verification steps"),
        ]
    )
    action_by_id = (
        {item.finding_id: item for item in report.ai_analysis.finding_actions}
        if report.ai_analysis
        else {}
    )
    for finding in report.findings:
        action = action_by_id.get(finding.id)
        title = localize_rule_field(finding.rule_id, locale, "title", finding.title)
        impact = localize_rule_field(finding.rule_id, locale, "impact", finding.impact)
        recommendation = localize_rule_field(
            finding.rule_id,
            locale,
            "recommendation",
            finding.recommendation,
        )
        line: int | str | None = finding.line_start
        if finding.line_start is not None and finding.line_end not in (None, finding.line_start):
            line = f"{finding.line_start}-{finding.line_end}"
        finding_sheet.append(
            [
                _cell(t(finding.severity.value.title())),
                _cell(finding.rule_id),
                _cell(title),
                _cell(finding.category),
                finding.confidence,
                _cell(finding.source),
                _cell(finding.page),
                _cell(line),
                _cell(finding.evidence_highlight),
                _cell(finding.evidence),
                _cell(impact),
                _cell(recommendation),
                _cell(action.remediation if action else ""),
                _cell("\n".join(action.verification_steps) if action else ""),
            ]
        )
    _style_sheet(finding_sheet, freeze="A2", auto_filter=True)

    if report.duplicate_groups:
        duplicate_sheet = workbook.create_sheet(t("Duplicate Comparison"))
        duplicate_sheet.append(
            [
                t("Category"),
                t("Similarity"),
                t("Estimated redundant tokens"),
                t("Reference occurrence"),
                t("File"),
                t("Page"),
                t("Line"),
                t("Compared excerpt"),
                t("Shared phrases"),
                t("Matched content"),
                t("Affected chunks"),
                t("Group count"),
            ]
        )
        for group in report.duplicate_groups:
            if not group.members:
                duplicate_sheet.append(
                    [
                        t(_duplicate_kind(group.category)),
                        group.similarity,
                        group.estimated_redundant_tokens,
                        "",
                        t(
                            "Comparison details are unavailable in this older report. Run the scan again."
                        ),
                        "",
                        "",
                        "",
                        "",
                        _cell(group.matched_content),
                        group.affected_chunks,
                        group.group_count,
                    ]
                )
                continue
            for member in group.members:
                duplicate_line: int | str | None = member.line_start
                if member.line_start is not None and member.line_end not in (
                    None,
                    member.line_start,
                ):
                    duplicate_line = f"{member.line_start}-{member.line_end}"
                duplicate_sheet.append(
                    [
                        _cell(t(_duplicate_kind(group.category))),
                        group.similarity,
                        group.estimated_redundant_tokens,
                        _cell(
                            t("Reference occurrence" if member.canonical else "Matching occurrence")
                        ),
                        _cell(member.source),
                        _cell(member.page),
                        _cell(duplicate_line),
                        _cell(member.evidence_excerpt),
                        _cell("\n".join(group.shared_phrases)),
                        _cell(group.matched_content),
                        group.affected_chunks,
                        group.group_count,
                    ]
                )
        _style_sheet(duplicate_sheet, freeze="A2", auto_filter=True)

    coverage = workbook.create_sheet(t("Coverage"))
    coverage.append([t("Area"), t("Status"), t("Reason")])
    for area, value in sorted(report.assessment_coverage.items()):
        coverage.append(
            [
                _cell(_coverage_area(area, t)),
                _cell(t(str(value.get("status")))),
                _cell(localize_coverage_reason(str(value.get("reason") or ""), locale)),
            ]
        )
    _style_sheet(coverage, freeze="A2", auto_filter=True)

    ingestion = workbook.create_sheet(t("Ingestion issues"))
    ingestion.append([t("Path"), t("Stage"), t("Issue"), t("Remediation")])
    for item in report.ingestion_issues:
        ingestion.append(
            [_cell(item.path), _cell(item.stage), _cell(item.message), _cell(item.remediation)]
        )
    _style_sheet(ingestion, freeze="A2", auto_filter=True)

    if report.rag_configuration_advice:
        rag = report.rag_configuration_advice
        rag_sheet = workbook.create_sheet(t("RAG Configuration"))
        rag_sheet.append([t("Area"), t("Report details")])
        _append_key_value(rag_sheet, t("Workload profile"), t(rag.profile.value))
        for key, value in rag.recommended.items():
            _append_key_value(
                rag_sheet,
                t(f"Recommended {key.replace('_', ' ')}"),
                t(value) if isinstance(value, str) else value,
            )
        for key, value in rag.observed.items():
            _append_key_value(rag_sheet, t(f"Observed {key.replace('_', ' ')}"), value)
        _append_key_value(rag_sheet, t("Actions"), "\n".join(t(item) for item in rag.actions))
        _append_key_value(
            rag_sheet,
            t("Validation metrics"),
            "\n".join(t(item) for item in rag.validation_metrics),
        )
        _append_key_value(
            rag_sheet, t("Limitations"), "\n".join(t(item) for item in rag.limitations)
        )
        _style_sheet(rag_sheet)

    if report.ai_analysis or report.ai_analysis_error:
        ai_sheet = workbook.create_sheet(t("AI Analysis"))
        ai_sheet.append([t("Area"), t("Report details")])
        if report.ai_analysis:
            ai = report.ai_analysis
            _append_key_value(ai_sheet, t("Executive summary"), ai.ai_analysis)
            _append_key_value(ai_sheet, t("Score commentary"), ai.score_commentary)
            _append_key_value(
                ai_sheet,
                t("Coverage caveat"),
                ai.coverage_caveat,
            )
            _append_key_value(
                ai_sheet,
                t("Root cause analysis"),
                "\n\n".join(_ai_root_cause_text(item, t) for item in ai.root_causes),
            )
            _append_key_value(
                ai_sheet,
                t("Priority actions"),
                "\n\n".join(
                    _ai_action_text(item, t)
                    for item in sorted(ai.priority_actions, key=lambda value: value.order)
                ),
            )
            _append_key_value(
                ai_sheet,
                t("Questions for review"),
                "\n\n".join(_ai_question_text(item, t) for item in ai.review_questions),
            )
            _append_key_value(ai_sheet, t("Verification steps"), "\n".join(ai.verification_steps))
            _append_key_value(ai_sheet, t("Limitations"), "\n".join(ai.limitations))
        else:
            _append_key_value(ai_sheet, t("AI analysis unavailable"), report.ai_analysis_error_code)
            _append_key_value(ai_sheet, t("Reason"), report.ai_analysis_error)
        _style_sheet(ai_sheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _contains_cjk(report: ReportDocument) -> bool:
    return bool(re.search(r"[\u3400-\u9fff]", report.model_dump_json()))


def _register_cjk_font() -> tuple[str, str] | None:
    name = "RAGScannerCJK"
    if name in pdfmetrics.getRegisteredFontNames():
        return name, name
    candidates = (
        Path("/System/Library/Fonts/STHeiti Medium.ttc"),
        Path("C:/Windows/Fonts/msyh.ttc"),
        Path("C:/Windows/Fonts/simsun.ttc"),
        Path("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"),
        Path("/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc"),
    )
    for candidate in candidates:
        if candidate.is_file():
            pdfmetrics.registerFont(TTFont(name, candidate, subfontIndex=0))
            return name, name
    return None


def _pdf_fonts(locale: str, report: ReportDocument) -> tuple[str, str]:
    if locale == "zh-CN" or _contains_cjk(report):
        registered = _register_cjk_font()
        if registered:
            return registered
        if "STSong-Light" not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
        return "STSong-Light", "STSong-Light"
    regular = "RAGScannerSans"
    bold = "RAGScannerSansBold"
    if regular not in pdfmetrics.getRegisteredFontNames():
        font_root = Path(reportlab.__file__).resolve().parent / "fonts"
        pdfmetrics.registerFont(TTFont(regular, font_root / "Vera.ttf"))
        pdfmetrics.registerFont(TTFont(bold, font_root / "VeraBd.ttf"))
    return regular, bold


def _render_pdf(report: ReportDocument, locale: str) -> bytes:
    locale = _locale(locale)
    t = _translator(locale)
    regular_font, bold_font = _pdf_fonts(locale, report)
    output = BytesIO()
    document = BaseDocTemplate(
        output,
        pagesize=A4,
        rightMargin=16 * mm,
        leftMargin=16 * mm,
        topMargin=20 * mm,
        bottomMargin=18 * mm,
        title=t("RAGScanner report"),
        author="RAGScanner",
    )
    frame = Frame(
        document.leftMargin, document.bottomMargin, document.width, document.height, id="content"
    )

    def page(canvas, doc) -> None:  # type: ignore[no-untyped-def]
        canvas.saveState()
        canvas.setFont(regular_font, 8)
        canvas.setFillColor(colors.HexColor("#64748B"))
        canvas.drawString(document.leftMargin, 10 * mm, "RAGScanner")
        canvas.drawRightString(A4[0] - document.rightMargin, 10 * mm, f"{t('Page')} {doc.page}")
        canvas.restoreState()

    document.addPageTemplates(PageTemplate(id="report", frames=[frame], onPage=page))
    styles = getSampleStyleSheet()
    body = ParagraphStyle(
        "body",
        parent=styles["BodyText"],
        fontName=regular_font,
        fontSize=9,
        leading=13,
        textColor=colors.HexColor("#24364B"),
        spaceAfter=6,
    )
    title = ParagraphStyle(
        "title",
        parent=styles["Title"],
        fontName=bold_font,
        fontSize=22,
        leading=27,
        textColor=colors.HexColor("#075B70"),
        alignment=TA_CENTER,
        spaceAfter=12,
    )
    h2 = ParagraphStyle(
        "h2",
        parent=styles["Heading2"],
        fontName=bold_font,
        fontSize=14,
        leading=18,
        textColor=colors.HexColor("#10233D"),
        spaceBefore=10,
        spaceAfter=7,
    )
    h3 = ParagraphStyle(
        "h3",
        parent=styles["Heading3"],
        fontName=bold_font,
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#075B70"),
        spaceBefore=6,
        spaceAfter=3,
    )
    small = ParagraphStyle(
        "small", parent=body, fontSize=7.5, leading=10, textColor=colors.HexColor("#5F6D7D")
    )
    score_label = ParagraphStyle(
        "score-label",
        parent=small,
        fontName=bold_font,
        fontSize=7,
        leading=9,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#10233D"),
    )
    score_text = ParagraphStyle(
        "score-value",
        parent=body,
        fontName=bold_font,
        fontSize=9.5,
        leading=11,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#10233D"),
    )
    table_header = ParagraphStyle(
        "table-header",
        parent=small,
        fontName=bold_font,
        textColor=colors.white,
    )

    def esc(value: Any) -> str:
        return html.escape(_display(value, fallback=t("Not assessed")), quote=True)

    story: list[Any] = [Paragraph(esc(t("RAGScanner report")), title)]
    metadata = Table(
        [
            [
                Paragraph(
                    f"<b>{esc(t('Source'))}</b><br/>{esc(report.scan.get('source_name'))}", body
                ),
                Paragraph(
                    f"<b>{esc(t('Status'))}</b><br/>{esc(t(str(report.scan.get('status'))))}",
                    body,
                ),
                Paragraph(f"<b>{esc(t('Generated'))}</b><br/>{esc(report.generated_at)}", body),
            ]
        ],
        colWidths=[document.width / 3] * 3,
    )
    metadata.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EAF5F6")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#BCD9DC")),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#BCD9DC")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    story.extend([metadata, Spacer(1, 7), Paragraph(esc(t("Executive summary")), h2)])
    score_data = [
        [
            Paragraph(esc(label), score_label)
            for label in (
                t("Overall score"),
                t("Security score"),
                t("Content quality"),
                t("Efficiency"),
            )
        ],
        [
            Paragraph(esc(report.scores.get(key)), score_text)
            for key in ("overall", "security", "knowledge_quality", "efficiency")
        ],
    ]
    scores = Table(score_data, colWidths=[document.width / 4] * 4)
    score_value = report.scores.get("overall")
    score_fill = {
        "healthy": "#DDF3E5",
        "warning": "#FFF1B8",
        "poor": "#FFD8B5",
        "critical": "#FFC7CE",
        "unassessed": "#E8EDF1",
    }[_score_band(score_value)]
    scores.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, 0), bold_font),
                ("FONTNAME", (0, 1), (-1, 1), bold_font),
                ("FONTSIZE", (0, 0), (-1, 0), 7.5),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EAF5F6")),
                ("BACKGROUND", (0, 1), (0, 1), colors.HexColor(score_fill)),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#BCD9DC")),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#BCD9DC")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(scores)
    severity_data = [
        [
            Paragraph(esc(label), table_header)
            for label in (
                t("Severity"),
                t("Critical"),
                t("High"),
                t("Medium"),
                t("Low"),
                t("Info"),
            )
        ],
        [
            Paragraph("", score_text),
            *[
                Paragraph(str(report.severity_summary.get(name, 0)), score_text)
                for name in ("critical", "high", "medium", "low", "info")
            ],
        ],
    ]
    severity = Table(severity_data, colWidths=[document.width / 6] * 6)
    severity.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, 0), bold_font),
                ("FONTNAME", (0, 1), (-1, 1), regular_font),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#10233D")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#BCD9DC")),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#BCD9DC")),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.extend([Spacer(1, 7), severity])

    if report.rag_configuration_advice:
        rag = report.rag_configuration_advice
        recommended = rag.recommended
        rag_why = t(str(recommended.get("why") or "Not assessed"))
        story.extend(
            [
                Paragraph(esc(t("RAG configuration advice")), h2),
                Paragraph(
                    f"<b>{esc(t('Workload profile'))}:</b> {esc(t(rag.profile.value))} · "
                    f"<b>{esc(t('Recommended chunk range'))}:</b> "
                    f"{esc(recommended.get('minimum_tokens'))}–{esc(recommended.get('maximum_tokens'))} "
                    f"{esc(t('tokens'))} ({esc(t('target'))} {esc(recommended.get('target_tokens'))}) · "
                    f"<b>{esc(t('Overlap'))}:</b> {esc(recommended.get('overlap_tokens'))} · "
                    f"<b>{esc(t('Retrieval top-k'))}:</b> {esc(recommended.get('retrieval_top_k'))}",
                    body,
                ),
                Paragraph(esc(t("Why this range?")), h3),
                Paragraph(esc(rag_why), body),
                Paragraph(esc(t("Actions")), h3),
                *[Paragraph(f"- {esc(t(item))}", body) for item in rag.actions],
                Paragraph(esc(t("Validation metrics")), h3),
                *[Paragraph(f"- {esc(t(item))}", body) for item in rag.validation_metrics],
                Paragraph(
                    esc(
                        t(
                            "Starting point only; validate with representative queries before production use."
                        )
                    ),
                    small,
                ),
            ]
        )

    if report.duplicate_groups:
        story.extend(
            [
                Paragraph(esc(t("Duplicate comparisons")), h2),
                Paragraph(
                    esc(
                        t(
                            "Review both locations together. A match is not an automatic deletion decision."
                        )
                    ),
                    small,
                ),
            ]
        )
        for group in report.duplicate_groups:
            story.append(
                Paragraph(
                    f"<b>{esc(t(_duplicate_kind(group.category)))}</b> · "
                    f"{esc(t('Similarity'))}: {group.similarity * 100:.1f}% · "
                    f"{esc(t('Affected chunks'))}: {group.affected_chunks} · "
                    f"{esc(t('Group count'))}: {group.group_count} · "
                    f"{esc(t('Estimated redundant tokens'))}: "
                    f"{group.estimated_redundant_tokens}",
                    h3,
                )
            )
            if group.matched_content:
                story.append(
                    Paragraph(
                        f"<b>{esc(t('Matched content'))}:</b> "
                        f'<font backColor="#FFF1A8">{esc(group.matched_content)}</font>',
                        body,
                    )
                )
            if group.shared_phrases:
                story.append(
                    Paragraph(
                        f"<b>{esc(t('Shared phrases'))}:</b> "
                        + " · ".join(esc(value) for value in group.shared_phrases),
                        body,
                    )
                )
            if not group.members:
                story.append(
                    Paragraph(
                        esc(
                            t(
                                "Comparison details are unavailable in this older report. Run the scan again."
                            )
                        ),
                        body,
                    )
                )
                continue
            for member in group.members[:_PDF_OCCURRENCES_PER_GROUP]:
                role = "Reference occurrence" if member.canonical else "Matching occurrence"
                location = _duplicate_member_location(member, t)
                excerpt = member.evidence_excerpt or t("Not assessed")
                story.extend(
                    [
                        Paragraph(
                            f"<b>{esc(t(role))}:</b> {esc(location)}",
                            body,
                        ),
                        Paragraph(
                            f'<font backColor="#FFF1A8">{esc(excerpt)}</font>',
                            small,
                        ),
                    ]
                )
            omitted = len(group.members) - _PDF_OCCURRENCES_PER_GROUP
            if omitted > 0:
                message = t(
                    "{count} more affected chunks omitted; use HTML or Excel for the complete list."
                ).format(count=omitted)
                story.append(Paragraph(esc(message), small))

    if report.ai_analysis:
        ai = report.ai_analysis
        story.extend(
            [
                Paragraph(esc(t("AI-assisted analysis")), h2),
                Paragraph(esc(ai.ai_analysis), body),
                Paragraph(esc(t("Score commentary")), h3),
                Paragraph(esc(ai.score_commentary), body),
            ]
        )
        if ai.coverage_caveat:
            story.append(
                Paragraph(
                    f"<i><b>{esc(t('Coverage caveat'))}:</b> {esc(ai.coverage_caveat)}</i>",
                    body,
                )
            )
        if ai.root_causes:
            story.append(Paragraph(esc(t("Root cause analysis")), h3))
            story.extend(
                Paragraph(
                    esc(_ai_root_cause_text(item, t)).replace("\n", "<br/>"),
                    body,
                )
                for item in ai.root_causes
            )
        if ai.priority_actions:
            story.append(Paragraph(esc(t("Priority actions")), h3))
            story.extend(
                Paragraph(
                    esc(_ai_action_text(item, t)).replace("\n", "<br/>"),
                    body,
                )
                for item in sorted(ai.priority_actions, key=lambda value: value.order)
            )
        if ai.review_questions:
            story.append(Paragraph(esc(t("Questions for review")), h3))
            story.extend(
                Paragraph(
                    esc(_ai_question_text(item, t)).replace("\n", "<br/>"),
                    body,
                )
                for item in ai.review_questions
            )
        story.append(
            Paragraph(f"{esc(ai.provider)} · {esc(ai.model)} · {esc(t(ai.disclaimer))}", small)
        )
    elif report.ai_analysis_error:
        story.extend(
            [
                Paragraph(esc(t("AI analysis unavailable")), h2),
                Paragraph(
                    f"<b>{esc(report.ai_analysis_error_code)}</b> {esc(report.ai_analysis_error)}",
                    body,
                ),
            ]
        )

    regular_findings = [
        finding for finding in report.findings if not finding.metadata.get("group_id")
    ]
    finding_heading = "Other findings" if report.duplicate_groups else "Findings"
    story.append(Paragraph(f"{esc(t(finding_heading))} ({len(regular_findings)})", h2))
    if not regular_findings:
        story.append(Paragraph(esc(t("No findings recorded.")), body))
    action_by_id = (
        {item.finding_id: item for item in report.ai_analysis.finding_actions}
        if report.ai_analysis
        else {}
    )
    grouped_findings: defaultdict[
        tuple[str, str, str, str, str, tuple[str, ...]], list[ReportFinding]
    ] = defaultdict(list)
    for finding in regular_findings:
        action = action_by_id.get(finding.id)
        remediation = action.remediation if action else finding.recommendation
        verification = tuple(action.verification_steps) if action else ()
        grouped_findings[
            (
                finding.rule_id,
                finding.title,
                finding.severity.value,
                finding.impact,
                remediation,
                verification,
            )
        ].append(finding)
    for index, (key, findings) in enumerate(grouped_findings.items(), start=1):
        rule_id, finding_title, severity_value, impact, remediation, verification = key
        localized_title = localize_rule_field(rule_id, locale, "title", finding_title)
        localized_impact = localize_rule_field(rule_id, locale, "impact", impact)
        localized_remediation = (
            remediation
            if any(finding.id in action_by_id for finding in findings)
            else localize_rule_field(rule_id, locale, "recommendation", remediation)
        )
        affected_chunks = len({finding.chunk_id or finding.id for finding in findings})
        group_ids = {
            str(finding.metadata["group_id"])
            for finding in findings
            if finding.metadata.get("group_id")
        }
        group_count = len(group_ids) or 1
        occurrence_label = (
            f" ({esc(t('Affected chunks'))}: {affected_chunks} · "
            f"{esc(t('Group count'))}: {group_count})"
            if len(findings) > 1
            else ""
        )
        blocks: list[Any] = [
            Paragraph(f"{index}. {esc(localized_title)}{occurrence_label}", h3),
            Paragraph(
                f"<b>{esc(t('Severity'))}:</b> {esc(t(severity_value.title()))} &nbsp; <b>{esc(t('Rule'))}:</b> {esc(rule_id)}",
                body,
            ),
            Paragraph(f"<b>{esc(t('Impact'))}:</b> {esc(localized_impact)}", body),
            Paragraph(
                f"<b>{esc(t('Recommendation'))}:</b> {esc(localized_remediation)}",
                body,
            ),
        ]
        blocks.extend(Paragraph(f"- {esc(step)}", body) for step in verification)
        story.append(KeepTogether(blocks))
        for occurrence, finding in enumerate(findings[:_PDF_OCCURRENCES_PER_GROUP], start=1):
            evidence = finding.evidence_highlight or finding.evidence
            evidence_text = f" - {esc(evidence)}" if evidence else ""
            story.append(
                Paragraph(
                    f"<b>{occurrence}.</b> {esc(_finding_location(finding, t))}{evidence_text}",
                    small,
                )
            )
        omitted = len(findings) - _PDF_OCCURRENCES_PER_GROUP
        if omitted > 0:
            message = t(
                "{count} more affected chunks omitted; use HTML or Excel for the complete list."
            ).format(count=omitted)
            story.append(Paragraph(esc(message), small))
        story.append(Spacer(1, 4))

    story.append(Paragraph(esc(t("Coverage")), h2))
    coverage_data = [
        [
            Paragraph(esc(t("Area")), table_header),
            Paragraph(esc(t("Status")), table_header),
            Paragraph(esc(t("Reason")), table_header),
        ]
    ]
    coverage_data.extend(
        [
            Paragraph(esc(_coverage_area(area, t)), small),
            Paragraph(esc(t(str(value.get("status")))), small),
            Paragraph(
                esc(localize_coverage_reason(str(value.get("reason") or ""), locale)),
                small,
            ),
        ]
        for area, value in sorted(report.assessment_coverage.items())
    )
    coverage = LongTable(
        coverage_data, colWidths=[45 * mm, 32 * mm, document.width - 77 * mm], repeatRows=1
    )
    coverage.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, 0), bold_font),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#10233D")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D5E0E5")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(coverage)
    if report.ingestion_issues:
        story.append(Paragraph(esc(t("Ingestion issues")), h2))
        for item in report.ingestion_issues:
            story.extend(
                [
                    Paragraph(f"<b>{esc(item.path)}</b> · {esc(item.stage)}", h3),
                    Paragraph(esc(item.message), body),
                    Paragraph(esc(item.remediation), small),
                ]
            )
    if report.methodology or report.limitations:
        story.append(Paragraph(esc(t("Methodology")), h2))
        story.extend(Paragraph(f"- {esc(t(item))}", body) for item in report.methodology)
        if report.limitations:
            story.append(Paragraph(esc(t("Limitations")), h3))
            story.extend(Paragraph(f"- {esc(t(item))}", body) for item in report.limitations)
    document.build(story)
    return output.getvalue()
