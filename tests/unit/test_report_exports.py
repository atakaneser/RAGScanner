"""Standalone dashboard report export tests."""

from io import BytesIO

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader
from ragscanner.quality import RAGConfigurationAdvice, RAGProfile
from ragscanner.reporting import export_report, report_export_filename
from ragscanner.reporting.models import ReportDuplicateGroup, ReportDuplicateMember


def test_html_export_is_localized_standalone_and_escaped(report, finding) -> None:  # type: ignore[no-untyped-def]
    unsafe = finding("html")
    unsafe.title = "<script>alert(1)</script>"
    unsafe.source = "politika-tr.md"
    unsafe.page = 2
    unsafe.evidence_highlight = "MFA'yı devre dışı bırak"
    exported = export_report(report("scan-html", findings=[unsafe]), "html", locale="tr")
    text = exported.content.decode("utf-8")

    assert exported.media_type == "text/html"
    assert exported.extension == "html"
    assert '<html lang="tr">' in text
    assert "Yönetici özeti" in text
    assert "Sorunlu metin" in text
    assert "<script>alert(1)</script>" not in text
    assert "&lt;script&gt;alert(1)&lt;/script&gt;" in text
    assert "connect-src 'none'" in text
    assert "<script" not in text.casefold()


def test_xlsx_export_has_structured_sheets_and_blocks_formulas(report, finding) -> None:  # type: ignore[no-untyped-def]
    suspicious = finding("xlsx")
    suspicious.title = '=HYPERLINK("https://example.invalid","open")'
    suspicious.source = "+SUM(A1:A2)"
    suspicious.evidence_highlight = "Şüpheli metin"
    exported = export_report(report("scan-xlsx", findings=[suspicious]), "xlsx", locale="tr")
    workbook = load_workbook(BytesIO(exported.content), data_only=False)

    assert exported.extension == "xlsx"
    assert workbook.sheetnames[:4] == ["Özet", "Bulgular", "Kapsam", "Veri alımı sorunları"]
    findings_sheet = workbook["Bulgular"]
    assert findings_sheet.freeze_panes == "A2"
    assert findings_sheet.auto_filter.ref
    assert findings_sheet["C2"].data_type == "s"
    assert str(findings_sheet["C2"].value).startswith("'=")
    assert str(findings_sheet["F2"].value).startswith("'+")
    assert findings_sheet["I2"].value == "Şüpheli metin"


def test_all_exports_include_rag_configuration_advice(report) -> None:  # type: ignore[no-untyped-def]
    value = report("scan-rag").model_copy(
        update={
            "rag_configuration_advice": RAGConfigurationAdvice(
                profile=RAGProfile.POLICY_PROCEDURE,
                configured={"target_tokens": 300},
                recommended={
                    "minimum_tokens": 80,
                    "target_tokens": 450,
                    "maximum_tokens": 700,
                    "overlap_tokens": 45,
                    "retrieval_top_k": 6,
                },
                observed={"median_chunk_tokens": 220},
                actions=["Benchmark representative policy questions."],
                validation_metrics=["Recall@k", "faithfulness"],
                limitations=["Starting point only."],
            )
        }
    )

    html_text = export_report(value, "html").content.decode("utf-8")
    workbook = load_workbook(BytesIO(export_report(value, "xlsx").content))
    pdf_text = "\n".join(
        page.extract_text() or ""
        for page in PdfReader(BytesIO(export_report(value, "pdf").content)).pages
    )

    assert "RAG configuration advice" in html_text
    assert "RAG Configuration" in workbook.sheetnames
    assert "RAG configuration advice" in pdf_text
    assert "450" in html_text


def test_all_exports_explain_duplicate_groups_with_both_locations(report) -> None:  # type: ignore[no-untyped-def]
    members = [
        ReportDuplicateMember(
            item_type="chunk",
            item_id=f"chunk-{index}",
            document_id=f"document-{index}",
            chunk_id=f"chunk-{index}",
            source=f"policy-{index}.md",
            page=index + 1,
            line_start=10,
            line_end=12,
            character_count=90,
            token_count=15,
            evidence_excerpt=f"Shared policy text from occurrence {index}.",
            canonical=index == 0,
        )
        for index in range(2)
    ]
    value = report("scan-duplicates").model_copy(
        update={
            "duplicate_groups": [
                ReportDuplicateGroup(
                    id="a" * 64,
                    category="near_duplicate_chunk",
                    canonical_item_id="chunk-0",
                    related_item_ids=["chunk-1"],
                    similarity=0.91,
                    estimated_redundant_characters=90,
                    estimated_redundant_tokens=15,
                    members=members,
                    shared_phrases=["shared policy text"],
                )
            ]
        }
    )

    html_text = export_report(value, "html", locale="tr").content.decode("utf-8")
    workbook = load_workbook(
        BytesIO(export_report(value, "xlsx", locale="tr").content), data_only=True
    )
    pdf_text = "\n".join(
        page.extract_text() or ""
        for page in PdfReader(BytesIO(export_report(value, "pdf", locale="tr").content)).pages
    )

    assert "Yinelenen içerik karşılaştırmaları" in html_text
    assert "policy-0.md" in html_text and "policy-1.md" in html_text
    assert "Shared policy text from occurrence 0." in html_text
    assert "Yinelenen Karşılaştırma" in workbook.sheetnames
    assert workbook["Yinelenen Karşılaştırma"]["E2"].value == "policy-0.md"
    assert "Yinelenen içerik karşılaştırmaları" in pdf_text
    assert "policy-0.md" in pdf_text and "policy-1.md" in pdf_text


@pytest.mark.parametrize(
    ("locale", "heading", "rerun"),
    [
        (
            "en",
            "Duplicate comparisons",
            "Comparison details are unavailable in this older report. Run the scan again.",
        ),
        (
            "tr",
            "Yinelenen içerik karşılaştırmaları",
            "Bu eski raporda karşılaştırma ayrıntıları yok. Taramayı yeniden çalıştırın.",
        ),
        (
            "de",
            "Duplikatvergleiche",
            "In diesem älteren Bericht fehlen Vergleichsdetails. Scan erneut ausführen.",
        ),
        (
            "fr",
            "Comparaisons des doublons",
            "Les détails de comparaison manquent dans cet ancien rapport. Relancez l’analyse.",
        ),
        ("zh-CN", "重复内容对比", "此旧报告没有对比详情，请重新运行扫描。"),
        (
            "it",
            "Confronti dei duplicati",
            "I dettagli non sono disponibili in questo vecchio rapporto. Esegui di nuovo la scansione.",
        ),
    ],
)
def test_historical_duplicate_guidance_is_localized(
    locale,
    heading,
    rerun,
    report,  # type: ignore[no-untyped-def]
) -> None:
    value = report("scan-old").model_copy(
        update={
            "duplicate_groups": [
                ReportDuplicateGroup(
                    id="b" * 64,
                    category="exact_duplicate_chunk",
                    canonical_item_id="chunk-0",
                    related_item_ids=["chunk-1"],
                    similarity=1,
                    estimated_redundant_characters=50,
                    estimated_redundant_tokens=8,
                )
            ]
        }
    )

    text = export_report(value, "html", locale=locale).content.decode("utf-8")

    assert heading in text
    assert rerun in text


@pytest.mark.parametrize("locale", ["en", "tr", "de", "fr", "zh-CN", "it"])
def test_pdf_export_is_readable_in_every_supported_locale(locale, report, finding) -> None:  # type: ignore[no-untyped-def]
    located = finding("pdf")
    located.source = "güvenlik-政策.pdf"
    located.page = 3
    located.evidence_highlight = "MFA güvenlik kontrolü"
    exported = export_report(report("scan-pdf", findings=[located]), "pdf", locale=locale)
    reader = PdfReader(BytesIO(exported.content))
    extracted = "\n".join(page.extract_text() or "" for page in reader.pages)

    assert exported.content.startswith(b"%PDF-")
    assert exported.media_type == "application/pdf"
    assert len(reader.pages) >= 1
    assert "RAGScanner" in extracted
    assert "scan-pdf" not in extracted  # exports do not expose an internal scan id as a heading


def test_pdf_groups_repeated_findings_and_bounds_occurrence_details(report, finding) -> None:  # type: ignore[no-untyped-def]
    repeated = []
    for index in range(80):
        item = finding("a")
        item.id = f"finding-{index}"
        item.fingerprint = f"{index:064x}"
        item.title = "Excessive Overlap"
        item.rule_id = "QUALITY-CHUNK-EXCESSIVE-OVERLAP"
        item.source = f"support-{index:02}.md"
        item.impact = "Poor chunk quality can reduce retrieval precision, waste context, or hide source structure."
        item.recommendation = (
            "Reduce bounded overlap without crossing unrelated structural boundaries."
        )
        repeated.append(item)
    exported = export_report(report("scan-grouped", findings=repeated), "pdf", locale="tr")
    reader = PdfReader(BytesIO(exported.content))
    extracted = "\n".join(page.extract_text() or "" for page in reader.pages)

    assert "Aşırı bindirme" in extracted
    assert "Tekrarlar: 80" in extracted
    assert "60 tekrar daha PDF özetinden çıkarıldı" in extracted
    assert len(reader.pages) < 10


def test_pdf_renders_plain_source_apostrophes_instead_of_html_entities(report, finding) -> None:  # type: ignore[no-untyped-def]
    item = finding("apostrophe")
    item.evidence = "## VPN'e bağlanma\nVPN'e bağlanmak için ',,sms' ekleyin."
    item.evidence_highlight = "VPN'e bağlanmak"
    exported = export_report(report("scan-apostrophe", findings=[item]), "pdf", locale="tr")
    extracted = "\n".join(
        page.extract_text() or "" for page in PdfReader(BytesIO(exported.content)).pages
    )

    assert "VPN'e" in extracted
    assert "&#x27;" not in extracted


def test_report_export_filename_is_bounded_and_safe(report) -> None:  # type: ignore[no-untyped-def]
    value = report("scan-name", source_name="Policy / ../../ unsafe")
    filename = report_export_filename(value, "abcdef0123456789", "xlsx")

    assert filename == "ragscanner-policy-unsafe-abcdef012345.xlsx"
    assert "/" not in filename


def test_unknown_export_format_is_rejected(report) -> None:  # type: ignore[no-untyped-def]
    with pytest.raises(ValueError, match="unsupported report export format"):
        export_report(report("scan-invalid"), "csv")  # type: ignore[arg-type]
